"""Shared styling so both workbooks look like the same tool made them."""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

NAVY   = "1F3864"
STEEL  = "D9E1F2"
GREY   = "F2F2F2"
AMBER  = "FFF2CC"

WHITE_BOLD = Font(bold=True, color="FFFFFF", size=12)
HDR        = Font(bold=True, size=10)
SMALL      = Font(size=9, color="595959")
SMALL_I    = Font(size=9, color="595959", italic=True)
BOLD_KEY   = Font(bold=True, size=11)

FILL_TITLE = PatternFill("solid", fgColor=NAVY)
FILL_HDR   = PatternFill("solid", fgColor=STEEL)
FILL_NOTE  = PatternFill("solid", fgColor=GREY)
FILL_KEY   = PatternFill("solid", fgColor=AMBER)
NOFILL     = PatternFill(fill_type=None)

_thin = Side(style="thin", color="BFBFBF")
BOX    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
NOBOX  = Border()

CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def scrub(ws, rows, cols):
    """Nuke values AND inherited formatting (fills/borders/fonts) in a block."""
    for r in rows:
        for c in cols:
            cell = ws.cell(r, c)
            cell.value = None
            cell.number_format = "General"
            cell.font = Font()
            cell.fill = NOFILL
            cell.border = NOBOX
            cell.alignment = Alignment()


def unmerge(ws):
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))


def title_bar(ws, text, sub, last_col_letter):
    """Navy title bar across row 1, optional grey subtitle on row 2.

    Pass sub=None when row 2 holds real data -- merging over a cell that another
    sheet references by address turns it into a MergedCell and silently feeds
    that sheet nothing. (Parameter!B2 is the module. Ask me how I know.)
    """
    ws.merge_cells("A1:%s1" % last_col_letter)
    c = ws["A1"]; c.value = text; c.font = WHITE_BOLD; c.fill = FILL_TITLE; c.alignment = LEFT
    ws.row_dimensions[1].height = 22
    if sub is not None:
        ws.merge_cells("A2:%s2" % last_col_letter)
        c = ws["A2"]; c.value = sub; c.font = SMALL_I; c.alignment = LEFT


def widths(ws, spec):
    for col, w in spec:
        ws.column_dimensions[col].width = w
