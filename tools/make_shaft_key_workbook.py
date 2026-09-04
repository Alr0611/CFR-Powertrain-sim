"""Build "CFR27 Shafts.xlsx" -- parallel-key and shaft sizing for the CFR27 driveline.

Deliberately mirrors the CBR27 (Baja) shaft workbook so the two read the same way:
same Notation / Unit / Formula / Comments / Value block layout, same method
(Mott, Machine Elements in Mechanical Design 6th ed.).

    Keys    Mott 11-4, eqs 11-1 .. 11-4. Key W and H looked up from the bore via
            Mott Table 11-1, so they update when a bore changes.
    Shafts  Mott 12-24 (fluctuating bending + steady torsion), with Mott's
            endurance-limit modifiers.

EVERYTHING IS A LIVE FORMULA. Type a real bore in and every downstream number moves.
Cells shaded YELLOW are geometry that has to come off the CAD/STEP; they are seeded
with placeholders so the sheet computes, and they are the only things to replace.

Regenerate:  python tools/make_shaft_key_workbook.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule

OUT = os.path.join(os.path.expanduser("~"), "Downloads", "CFR27 Shafts.xlsx")
EQ  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "eq_img_keys")

def eq(ws, cell, name, height=52):
    """Drop a rendered Mott equation into the sheet at `cell`."""
    f = os.path.join(EQ, name + ".png")
    if not os.path.exists(f):
        ws[cell] = f"[missing {name}.png -- run tools/make_key_eq_images.py]"
        return
    img = XLImage(f)
    img.height = height
    img.width = int(img.width * height / XLImage(f).height)
    ws.add_image(img, cell)

# ---------------------------------------------------------------- styling
H1     = Font(bold=True, size=14)
H2     = Font(bold=True, size=11)
HDR    = Font(bold=True, italic=True, size=9)
MONO   = Font(name="Consolas", size=10)
WARN   = Font(bold=True, color="9C0006")
GREY   = Font(italic=True, size=9, color="595959")

F_CAD   = PatternFill("solid", fgColor="FFF2CC")   # yellow: FROM CAD, replace me
F_CALC  = PatternFill("solid", fgColor="E2EFDA")   # green: computed, do not type here
F_MEAS  = PatternFill("solid", fgColor="DDEBF7")   # blue: MEASURED off a STEP file
F_HDR   = PatternFill("solid", fgColor="D9E1F2")
F_TITLE = PatternFill("solid", fgColor="BDD7EE")
F_BAD   = PatternFill("solid", fgColor="FFC7CE")

THIN = Side(style="thin", color="BFBFBF")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def block_header(ws, row):
    for col, txt in zip("BCDE", ["Notation", "Unit", "Formula", "Comments"]):
        c = ws[f"{col}{row}"]
        c.value, c.font, c.fill = txt, HDR, F_HDR


def put(ws, row, label, notation="", unit="", formula="", comment="", value=None,
        fill=None, num="0.000"):
    """One line of a CBR-style block. Value lands in column F."""
    ws[f"A{row}"] = label
    ws[f"B{row}"] = notation
    ws[f"C{row}"] = unit
    ws[f"D{row}"] = formula
    ws[f"E{row}"] = comment
    ws[f"D{row}"].font = MONO
    ws[f"E{row}"].font = GREY
    if value is not None:
        c = ws[f"F{row}"]
        c.value = value
        c.number_format = num
        c.border = BOX
        if fill:
            c.fill = fill
    return row + 1


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


wb = Workbook()

# ================================================================== READ ME
ws = wb.active
ws.title = "Read Me"
widths(ws, {"A": 108})
rows = [
    ("CFR27 SHAFT AND KEY CALCULATIONS", H1, None),
    ("", None, None),
    ("WHAT THIS IS", H2, None),
    ("The CFR27 shafts are plain round right now, so only a press/clamp fit is carrying torque into", None, None),
    ("the gears and sprockets. This workbook sizes the parallel keys that replace that, and checks the", None, None),
    ("shafts those keyseats are cut into. Built to match the CBR27 (Baja) shaft workbook block for", None, None),
    ("block, so the two can be read side by side and reviewed the same way.", None, None),
    ("", None, None),
    ("METHOD -- Mott, Machine Elements in Mechanical Design, 6th ed.", H2, None),
    ("  Keys, section 11-4:", None, None),
    ("      shear over the W x L section        tau   = 2T/(D*W*L)                      eq 11-1", None, MONO),
    ("      design shear stress, MSST           tau_d = 0.5*sy/N", None, MONO),
    ("      min length, shear                   Lmin  = 2T/(tau_d*D*W)                  eq 11-2", None, MONO),
    ("      bearing on the L x H/2 flank        sigma = 4T/(D*L*H)                      eq 11-3", None, MONO),
    ("      design bearing stress               sig_d = sy/N", None, MONO),
    ("      min length, bearing                 Lmin  = 4T/(sig_d*D*H)                  eq 11-4", None, MONO),
    ("  Key size W x H is NOT a free choice: Mott Table 11-1 fixes it from the shaft diameter", None, None),
    ("  (sheet 'Key Size Table', looked up live). Only LENGTH and MATERIAL are design variables.", None, None),
    ("", None, None),
    ("  Shafts, section 12-8, eq 12-24 (fluctuating bending + steady torsion):", None, None),
    ("      D = [ (32N/pi) * sqrt( (Kt*M/sn)^2 + (3/4)*(T/sy)^2 ) ]^(1/3)", None, MONO),
    ("  with the endurance limit modified per Mott ch.5: sn = sn' * Cs * CR * Cm.", None, None),
    ("", None, None),
    ("THE SHEETS", H2, None),
    ("  Key Spec (MAKE THIS)  the answer: width, height, keyseat depths and minimum key length per joint,", None, None),
    ("                        with the order length for each DIN 6885 end form. Start here.", None, None),
    ("  Formulas              every equation used, rendered, each citing its Mott number.", None, None),
    ("  Inputs                torque, ratios, materials, design factor. Everything else references it.", None, None),
    ("  Key Size Table        Mott Table 11-1 + DIN 6885-1 keyseat depths and preferred key lengths.", None, None),
    ("  Keys                  the full Mott 11-4 working, one block per joint.", None, None),
    ("  Geometry (from STEP)  every dimension measured off the CAD, and what is still missing.", None, None),
    ("  Shafts                Mott 12-24 shaft sizing, plus the static torsion screen.", None, None),
    ("", None, None),
    ("THE ONE THING TO UNDERSTAND ABOUT KEY LENGTH", H2, None),
    ("Mott's equations give a BEARING length -- how much key has to be in contact. What you ORDER is", None, None),
    ("longer than that if the key has radiused ends, because a radiused end does not bear over its own", None, None),
    ("projection. DIN 6885-1 Form A is round both ends (add W), Form AB is one round end (add W/2),", None, None),
    ("Form B is square both ends (add nothing). So for the same torque a square-ended key is the SHORTEST", None, None),
    ("part you can make. Both CFR27 keys are currently Form AB.", None, None),
    ("The catch: a square-ended key cannot sit in the semicircular ends of an end-milled slot, so the most", None, None),
    ("that fits is the slot's straight portion, i.e. slot length minus one cutter diameter. The Key Spec", None, None),
    ("sheet works that out per joint and says whether the end form alone is enough to save the joint.", None, None),
    ("", None, None),
    ("HOW TO USE IT", H2, None),
    ("  1. YELLOW cells are geometry that must come off the CAD / STEP files. They are seeded with", None, None),
    ("     placeholders purely so the sheet computes. Replace every one of them.", None, None),
    ("  2. GREEN cells are computed. Do not type in them.", None, None),
    ("  3. Everything else follows automatically -- change a bore and the key size, both minimum", None, None),
    ("     lengths, the factor of safety and the fits/does-not-fit verdict all move with it.", None, None),
    ("", None, None),
    ("WHAT IS ALREADY REAL, AND WHAT IS NOT", H2, None),
    ("  REAL       every torque. Motor cap 150 Nm and the measured 123 Nm VC ceiling from", None, None),
    ("             params_cfr26.m; the 15/30 spur and 13/30 chain from the actual car; the grip", None, None),
    ("             cap from the measured Hoosier tyre model. None of it is invented.", None, None),
    ("  PLACEHOLDER  every bore, hub length, bearing span and pitch diameter (all YELLOW).", None, None),
    ("  ASSUMED    the shaft, hub and key material grades. Confirm the stock and heat treat.", None, None),
    ("", None, None),
    ("CROSS-CHECK", H2, None),
    ("  shaft_key_calc.m in the CFR-Powertrain repo runs the same equations from params_cfr26.m", None, None),
    ("  and plots required key length against bore as a continuous sweep. Use it to see how far a", None, None),
    ("  bore has to grow before a key fits; use this workbook to write the drawing and justify it.", None, None),
    ("", None, None),
    ("SOURCES", H2, None),
    ("  Mott, Machine Elements in Mechanical Design, 6th ed., ch.11 (keys) and ch.12 (shafts).", None, None),
    ("  Mott Table 11-1 key sizes = ISO/R 773 = DIN 6885-1, so these are stock keys, not specials.", None, None),
    ("  Mott Table 11-4 material strengths.", None, None),
    ("  Torques, ratios, grip: CFR-Powertrain/params_cfr26.m + gear_decision_summary.m.", None, None),
]
for i, (txt, font, alt) in enumerate(rows, start=1):
    c = ws.cell(row=i, column=1, value=txt)
    if font:
        c.font = font
    elif alt:
        c.font = alt

# ================================================================== INPUTS
ws = wb.create_sheet("Inputs")
widths(ws, {"A": 46, "B": 12, "C": 8, "D": 34, "E": 46, "F": 14})
ws["A1"] = "DESIGN INPUTS -- every other sheet references this one"
ws["A1"].font = H1
ws["A1"].fill = F_TITLE

r = 3
ws[f"A{r}"] = "Torque source"; ws[f"A{r}"].font = H2; r += 1
block_header(ws, r); r += 1
r = put(ws, r, "Motor torque, DESIGN case", "T design", "Nm", "",
        "*** THIS IS THE NUMBER EVERYTHING IS SIZED ON. *** Set ABOVE the datasheet peak on purpose: "
        "the EMRAX can transiently exceed its rated peak, the torque map can be changed in firmware next "
        "season, and nobody wants to find out by shearing a key. Dial it here and the whole workbook re-solves.",
        170.0, F_CAD, "0.0")
r_Tds = r
r = put(ws, r, "Motor torque, datasheet peak", "T peak", "Nm", "params_cfr26 p.T_flat_cap",
        "EMRAX 208 rated peak. Reference only -- the design case above is what is used.", 150.0, F_CALC, "0.0")
r_Tdrv = r
r = put(ws, r, "Motor torque, as driven", "T drv", "Nm", "params_cfr26 p.T_driver_max",
        "MEASURED VC ceiling off the log -- what the car actually asks for today.", 123.0, F_CALC, "0.0")
r_ovl = r
r = put(ws, r, "Implied overload factor", "k ovl", "ul", f"=F5/F{r_Tds}",
        "Design torque over datasheet peak. This is a LOAD-uncertainty allowance and it is separate from "
        "the design factor N below, which covers MATERIAL and MACHINING uncertainty. Stacking the two is "
        "normal practice (it is what a service factor is), but be aware the effective margin against yield "
        "is k_ovl x N -- see the row below so it is never a surprise.",
        f"=F5/F{r_Tds}", F_CALC, "0.000")
r += 1

ws[f"A{r}"] = "Driveline"; ws[f"A{r}"].font = H2; r += 1
block_header(ws, r); r += 1
r_spur_dr = r
r = put(ws, r, "Spur driver teeth (motor shaft)", "z1", "T", "", "Fixed unless the housing changes", 15, F_CAD, "0")
r_spur_dn = r
r = put(ws, r, "Spur driven teeth", "z2", "T", "", "", 30, F_CAD, "0")
r_spur = r
r = put(ws, r, "Spur ratio", "i spur", "ul", f"=F{r_spur_dn}/F{r_spur_dr}", "", f"=F{r_spur_dn}/F{r_spur_dr}", F_CALC, "0.0000")
r_ch_dr = r
r = put(ws, r, "Chain driver teeth (intermediate)", "z3", "T", "", "THIS is the gear-ratio knob", 13, F_CAD, "0")
r_ch_dn = r
r = put(ws, r, "Chain driven teeth (diff)", "z4", "T", "", "", 30, F_CAD, "0")
r_chain = r
r = put(ws, r, "Chain ratio", "i chain", "ul", f"=F{r_ch_dn}/F{r_ch_dr}", "", f"=F{r_ch_dn}/F{r_ch_dr}", F_CALC, "0.0000")
r_G = r
r = put(ws, r, "Total gear ratio", "G", "ul", f"=F{r_spur}*F{r_chain}", "Currently 4.6154", f"=F{r_spur}*F{r_chain}", F_CALC, "0.0000")
r_eta = r
r = put(ws, r, "Drivetrain efficiency", "eta", "ul", "params_cfr26 p.eta_drivetrain",
        "12 deg halfshaft stack. Only used for the grip comparison below, NOT for key torque (a key sees the torque on ITS shaft, before the losses downstream of it).",
        0.794, F_CALC, "0.000")
r_open = r
r = put(ws, r, "Open diff? (1 = open, 0 = spool/locked)", "", "ul", "",
        "An OPEN diff splits torque evenly, so each output stub sees G*T/2. A SPOOL or a locked LSD can put the LOT through one side with the inside wheel light. Set 0 and the stub key is sized on the full G*T.",
        1, F_CAD, "0")
r += 1

ws[f"A{r}"] = "Materials (Mott Table 11-4)"; ws[f"A{r}"].font = H2; r += 1
block_header(ws, r); r += 1
r_su_key = r
r = put(ws, r, "Ultimate strength, key", "Su", "MPa", "", "SAE 1045", 627, F_CAD, "0")
r_sy_key = r
r = put(ws, r, "Yield strength, KEY material", "sy key", "MPa", "", "SAE 1045. Mott's default is 1018 (weaker, sacrificial). Stepped up because a sheared key mid-endurance is a DNF and the gear it protects is re-machinable over the winter.", 531, F_CAD, "0")
r_sy_sh = r
r = put(ws, r, "Yield strength, SHAFT material", "sy shaft", "MPa", "", "*** ASSUMED 4140. CONFIRM STOCK + HEAT TREAT ***", 621, F_CAD, "0")
r_su_sh = r
r = put(ws, r, "Ultimate strength, SHAFT material", "Su shaft", "MPa", "", "*** ASSUMED 4140 ***", 703, F_CAD, "0")
r_sy_hub = r
r = put(ws, r, "Yield strength, HUB (gear/sprocket)", "sy hub", "MPa", "",
        "MEASURED FROM CAD: the gears are KHK MSGA2.5-15 and MSGA2.5-30, stock ground spur gears in "
        "SCM415 carburised. Core yield taken as 800 MPa, the same figure the CBR27 Baja sheet used for "
        "SCM415. Note this is now the STRONGEST of the three bearing surfaces, so it no longer limits "
        "anything -- the shaft does.",
        800, F_MEAS, "0")
r_sy_bear = r
r = put(ws, r, "Governing bearing yield", "sy bear", "MPa",
        f"=MIN(F{r_sy_key},F{r_sy_sh},F{r_sy_hub})",
        "Mott 11-4: bearing fails on whichever of key / shaft keyseat / hub keyseat has the LOWEST compressive yield. Not necessarily the key.",
        f"=MIN(F{r_sy_key},F{r_sy_sh},F{r_sy_hub})", F_CALC, "0")
r += 1

ws[f"A{r}"] = "Design stresses"; ws[f"A{r}"].font = H2; r += 1
block_header(ws, r); r += 1
r_N = r
r = put(ws, r, "Design factor", "N", "ul", "", "CBR27 ran 1.25-2.0. Torque here is a measured cap on a measured car, so the uncertainty is material and machining, not load. 2.0 = the conservative end.", 2.0, F_CAD, "0.00")
r_eff = r
r = put(ws, r, "EFFECTIVE margin vs the datasheet peak", "k ovl x N", "ul", f"=F{r_ovl}*F{r_N}",
        "What the parts are really being held to, relative to the 150 Nm datasheet peak. Sanity-check this "
        "is what you meant: if it looks too rich, lower the design torque rather than N -- N is the one "
        "protecting you against the material grade, which is still unconfirmed.",
        f"=F{r_ovl}*F{r_N}", F_CALC, "0.00")
r_taud = r
r = put(ws, r, "Design shearing strength", "tau d", "MPa", f"tau d = 0.5*sy/N",
        "MSST theory of failure in shear", f"=0.5*F{r_sy_key}/F{r_N}", F_CALC, "0.00")
r_sigd = r
r = put(ws, r, "Design compressive strength", "sigma d", "MPa", f"sigma d = sy bear/N",
        "", f"=F{r_sy_bear}/F{r_N}", F_CALC, "0.00")
r_phi = r
r = put(ws, r, "Gear pressure angle", "phi", "deg", "", "*** FROM THE GEAR DRAWING. 20 deg assumed. *** This is one of the two numbers gear_check.m has been parked waiting on.", 20, F_CAD, "0")
r_Kt = r
r = put(ws, r, "Fatigue stress conc., keyseat (bending)", "Kf", "ul", "Hindhede Table 15-4",
        "Profile (end-milled) keyseat in QUENCHED AND TEMPERED steel. From Hindhede, Machine Design "
        "Fundamentals, Table 15-4, after Lipson & Juvinall. These are Kf, NOT Kt, so no notch-sensitivity "
        "correction is applied. Other values: profile keyseat 1.6 if annealed; sled-runner 1.6 (Q&T) or "
        "1.3 (annealed). In TORSION the same table gives 1.6 (Q&T) / 1.3 (annealed) -- not used here "
        "because the torque is treated as steady, per Mott eq 12-24.",
        2.0, F_CAD, "0.0")
r += 1

ws[f"A{r}"] = "Grip cap (sanity check, not a design input)"; ws[f"A{r}"].font = H2; r += 1
block_header(ws, r); r += 1
r_Tgrip = r
r = put(ws, r, "Wheel torque the tyres allow", "T grip", "Nm", "mu * Fz_rear * r_load",
        "mu 1.43, rear Fz 2155 N at a 1.3 g launch, r_load 0.1901 m. From the measured Hoosier model in params_cfr26.",
        585.3, F_CALC, "0.0")
r_Tmw = r
r = put(ws, r, "Wheel torque if motor-limited", "T wheel", "Nm", f"=F5*F{r_G}*F{r_eta}",
        "", f"=F5*F{r_G}*F{r_eta}", F_CALC, "0.0")
r = put(ws, r, "Which governs at the wheel?", "", "", f"=IF(F{r_Tgrip}<F{r_Tmw},...)",
        "The grip cap is a STEADY cap. It does not cover wheel hop, a kerb strike, or dropping onto a locked wheel, so the keys are NOT trimmed down to it -- that is what N is absorbing.",
        f'=IF(F{r_Tgrip}<F{r_Tmw},"GRIP governs","MOTOR governs")', F_CALC, "General")

INP = "Inputs"
T_MOTOR   = f"{INP}!$F$5"
SPUR      = f"{INP}!$F${r_spur}"
GTOT      = f"{INP}!$F${r_G}"
OPENDIFF  = f"{INP}!$F${r_open}"
TAUD      = f"{INP}!$F${r_taud}"
SIGD      = f"{INP}!$F${r_sigd}"
SYKEY     = f"{INP}!$F${r_sy_key}"
SYBEAR    = f"{INP}!$F${r_sy_bear}"
SYSHAFT   = f"{INP}!$F${r_sy_sh}"
SUSHAFT   = f"{INP}!$F${r_su_sh}"
NDES      = f"{INP}!$F${r_N}"
PHI       = f"{INP}!$F${r_phi}"
KT        = f"{INP}!$F${r_Kt}"

# ================================================================== KEY SIZE TABLE
ws = wb.create_sheet("Key Size Table")
widths(ws, {"A": 16, "B": 18, "C": 12, "D": 12, "E": 14, "F": 60})
ws["A1"] = "MOTT TABLE 11-1  --  Key Size vs. Shaft Diameter (SI metric)"
ws["A1"].font = H1
ws["A1"].fill = F_TITLE
ws["A2"] = "Same table as ISO/R 773 and DIN 6885-1. Keys to these sizes are stock parts."
ws["A2"].font = GREY
ws["A3"] = "Looked up live by the Keys sheet. Do not reorder the rows -- the lookup is a sorted MATCH."
ws["A3"].font = WARN
for col, txt in zip("ABCDEFG", ["Over (mm)", "To-including (mm)", "Width W (mm)", "Height H (mm)",
                                "Section", "Shaft depth t1", "Hub depth t2"]):
    c = ws[f"{col}5"]
    c.value, c.font, c.fill = txt, HDR, F_HDR

# t1 (shaft keyseat depth) and t2 (hub keyseat depth) are DIN 6885-1. Mott's Table 11-1
# does not carry them, but a drawing cannot be made without them, and both CFR27 shafts
# were verified against this column: D15 measured 2.93 vs t1 3.0, D22 measured 3.46 vs 3.5.
TBL = [(6,8,2,2,1.2,1.0),(8,10,3,3,1.8,1.4),(10,12,4,4,2.5,1.8),(12,17,5,5,3.0,2.3),
       (17,22,6,6,3.5,2.8),(22,30,8,7,4.0,3.3),(30,38,10,8,5.0,3.3),(38,44,12,8,5.0,3.3),
       (44,50,14,9,5.5,3.8),(50,58,16,10,6.0,4.3),(58,65,18,11,7.0,4.4),(65,75,20,12,7.5,4.9),
       (75,85,22,14,9.0,5.4),(85,95,25,14,9.0,5.4),(95,110,28,16,10.0,6.4),
       (110,130,32,18,11.0,7.4),(130,150,36,20,12.0,8.4),(150,170,40,22,13.0,9.4),
       (170,200,45,25,15.0,10.4),(200,230,50,28,17.0,11.4),(230,260,56,32,20.0,12.4),
       (260,290,63,32,20.0,12.4),(290,330,70,36,22.0,14.4),(330,380,80,40,25.0,15.4),
       (380,440,90,45,28.0,17.4),(440,500,100,50,31.0,19.5)]
for i, (o, t, w, h, t1, t2) in enumerate(TBL):
    rr = 6 + i
    ws[f"A{rr}"], ws[f"B{rr}"], ws[f"C{rr}"], ws[f"D{rr}"] = o, t, w, h
    ws[f"E{rr}"] = "square" if w == h else "rectangular"
    ws[f"F{rr}"], ws[f"G{rr}"] = t1, t2
    for col in "ABCDEFG":
        ws[f"{col}{rr}"].border = BOX
FIRST, LAST = 6, 6 + len(TBL) - 1
ws[f"F{LAST+2}"] = ("Lookup used on the Keys sheet, and the -0.0001 is not a typo: the rows are "
                    "'over X to-including Y', so a bore of exactly 30 mm belongs to the 22-30 row "
                    "(W=8), not the 30-38 row (W=10). Nudging the lookup below the boundary lands "
                    "it on the right side.")
ws[f"F{LAST+2}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[LAST+2].height = 60

# DIN 6885-1 preferred key lengths -- a key is ordered off this list, not cut to an
# arbitrary number, so the spec sheet rounds UP to one of these.
ws["I5"] = "DIN 6885-1 preferred key lengths (mm)"
ws["I5"].font = HDR
ws["I5"].fill = F_HDR
PREF = [6,8,10,12,14,16,18,20,22,25,28,32,36,40,45,50,56,63,70,80,90,100,110,125,140,160,180,200]
for i, v in enumerate(PREF):
    c = ws.cell(row=6 + i, column=9, value=v)
    c.border = BOX
widths(ws, {"I": 32})
PFIRST, PLAST = 6, 6 + len(PREF) - 1
KTBL_PREF = f"'Key Size Table'!$I${PFIRST}:$I${PLAST}"
KTBL_T1   = f"'Key Size Table'!$F${FIRST}:$F${LAST}"
KTBL_T2   = f"'Key Size Table'!$G${FIRST}:$G${LAST}"

KTBL_OVER = f"'Key Size Table'!$A${FIRST}:$A${LAST}"
KTBL_W    = f"'Key Size Table'!$C${FIRST}:$C${LAST}"
KTBL_H    = f"'Key Size Table'!$D${FIRST}:$D${LAST}"

# ================================================================== KEYS
ws = wb.create_sheet("Keys")
widths(ws, {"A": 42, "B": 20, "C": 8, "D": 30, "E": 62, "F": 14})
ws["A1"] = "CFR27 PARALLEL KEY CALCULATIONS  --  Mott section 11-4"
ws["A1"].font = H1
ws["A1"].fill = F_TITLE
ws["A2"] = ("Sized on the DESIGN torque from the Inputs sheet, which is deliberately above the datasheet peak. "
            "YELLOW = from CAD, replace. BLUE = measured off a STEP file. GREEN = computed, do not type. "
            "W and H are looked up from the bore, so a bore change updates everything. "
            "See the Formulas sheet for every equation, and Key Spec for what to actually cut.")
ws["A2"].font = GREY

# station: (name, torque formula, bore, key length, comment, measured?)
# Bores and key lengths for the first two are MEASURED off the STEP files
# (motor_to_gear_shaft.STEP, shaft_to_sprocket.STEP) -- see the Geometry sheet.
STATIONS = [
    ("Shaft 1 (motor) / 15T spur pinion", f"={T_MOTOR}*1000", 15.0, 25.0,
     "MEASURED, bore 15.00. Key part Shaftkey_motor_to_gear.STEP is 5.00 x 5.00 x 27.50 long with one R2.50 end, so the STRAIGHT bearing length is 25.00 mm -- that is the L used here (the conservative convention: a radiused end does not bear over its full projection). Keyseat slot is 35.0 mm overall, so a longer key fits without touching the shaft.",
     True),
    ("Shaft 2 (interm.) / 30T spur gear", f"={T_MOTOR}*{SPUR}*1000", 22.0, 24.0,
     "MEASURED, bore 22.00. Key part shaftkey_gear_to_sprocket.STEP is 6.00 x 6.00 x 27.00 long with one R3.00 end, so the STRAIGHT bearing length is 24.00 mm. Keyseat slot is 30.0 mm overall, so the most this bore can ever reach is about 27 mm straight.",
     True),
    ("Diff input / 30T chain sprocket", f"={T_MOTOR}*{GTOT}*1000", 35, 30,
     "*** NO STEP FILE YET -- PLACEHOLDER *** Highest torque in the driveline, so this is likely the station that sets the bore. Often bolted to the diff carrier rather than keyed -- confirm.",
     False),
    ("Diff output stub / CV cup (per side)", f"={T_MOTOR}*{GTOT}*1000/IF({OPENDIFF}=1,2,1)", 25, 30,
     "*** NO STEP FILE YET -- PLACEHOLDER *** Halves across an OPEN diff. Set Inputs 'Open diff?' to 0 for a spool/locked LSD and this doubles.",
     False),
]

r = 4
key_rows = []
for name, tform, bore, hub, note, measured in STATIONS:
    geo_fill = F_MEAS if measured else F_CAD
    c = ws[f"A{r}"]
    c.value, c.font, c.fill = name, H2, F_TITLE
    r += 1
    block_header(ws, r)
    r += 1
    rT = r
    r = put(ws, r, "Torque at this joint", "T", "Nmm", tform.replace("=", "", 1), note, tform, F_CALC, "0")
    rD = r
    r = put(ws, r, "Bore / shaft diameter", "D", "mm", "",
            "MEASURED off the STEP file" if measured else "*** FROM CAD -- PLACEHOLDER ***",
            bore, geo_fill, "0.00")
    rW = r
    r = put(ws, r, "Keyway width", "W", "mm", "Mott Table 11-1, from D",
            "Not a free choice. It steps with the bore.",
            f"=INDEX({KTBL_W},MATCH(F{rD}-0.0001,{KTBL_OVER},1))", F_CALC, "0.0")
    rH = r
    r = put(ws, r, "Keyway height", "H", "mm", "Mott Table 11-1, from D", "",
            f"=INDEX({KTBL_H},MATCH(F{rD}-0.0001,{KTBL_OVER},1))", F_CALC, "0.0")
    rL = r
    r = put(ws, r, "Available key length", "Lmax", "mm", "",
            ("MEASURED off the KEY part: its straight (full-height) bearing length. The radiused end is excluded, which is the conservative convention. Still to confirm: the GEAR HUB length, since the key also cannot be longer than the hub it drives."
             if measured else
             "*** FROM CAD -- PLACEHOLDER *** Mott: run the key over a substantial part of the hub, but keep axial clearance from shoulder fillets and ring grooves so the stress raisers do not compound."),
            hub, geo_fill, "0.0")
    rtd = r
    r = put(ws, r, "Design shearing strength", "tau d", "MPa", "tau d = 0.5*sy/N",
            "MSST theory of failure in shear", f"={TAUD}", F_CALC, "0.00")
    rLs = r
    r = put(ws, r, "Min. length for shear", "Lmin (shear)", "mm", "Lmin = 2*T/(tau d*D*W)",
            "Mott eq 11-2", f"=2*F{rT}/(F{rtd}*F{rD}*F{rW})", F_CALC, "0.00")
    rsd = r
    r = put(ws, r, "Design compressive strength", "sigma d", "MPa", "sigma d = sy bear/N",
            "Against the WEAKEST of key / shaft / hub", f"={SIGD}", F_CALC, "0.00")
    rLb = r
    r = put(ws, r, "Min. length for compression", "Lmin (compression)", "mm",
            "Lmin = (4*T)/(sigma d*D*H)", "Mott eq 11-4", f"=4*F{rT}/(F{rsd}*F{rD}*F{rH})", F_CALC, "0.00")
    rLr = r
    r = put(ws, r, "GOVERNING minimum length", "Lmin", "mm", "= MAX of the two above",
            "This is the number that goes on the drawing.", f"=MAX(F{rLs},F{rLb})", F_CALC, "0.00")
    rN = r
    r = put(ws, r, "FOS if the key runs the full hub", "N actual", "ul",
            "Mott 11-2/11-4 solved for N",
            "What you actually get if the key is made as long as the hub allows.",
            f"=MIN(0.5*{SYKEY}*F{rD}*F{rW}*F{rL}/(2*F{rT}), {SYBEAR}*F{rD}*F{rH}*F{rL}/(4*F{rT}))",
            F_CALC, "0.00")
    rFit = r
    r = put(ws, r, "Does it fit?", "", "", f"=IF(Lmin<=Lmax)",
            "If NO, the fix is a BIGGER BORE, not a longer key: W and H grow with D, so Lmin falls roughly as D^2.",
            f'=IF(F{rLr}<=F{rL},"YES","NO - INCREASE BORE")', None, "General")
    ws[f"F{rFit}"].font = Font(bold=True)
    key_rows.append((name, rT, rD, rW, rH, rL, rLs, rLb, rLr, rN, rFit))
    r += 1

# --- the two joints on these shafts that are NOT keys, so nobody assumes they were checked
ws[f"A{r}"] = "JOINTS ON THESE SHAFTS THAT ARE NOT KEYS (not covered above)"
ws[f"A{r}"].font = H2
ws[f"A{r}"].fill = F_BAD
r += 1
for txt in [
    "Motor flange -> Shaft 1:  BOLTED, not keyed. Ø94 x 7 mm flange, 12 x Ø9 holes on a Ø75 bolt circle "
    "(M8 clearance), Ø56 pilot boss. Carries the full 150 Nm as a bolted friction/shear joint -- needs a "
    "bolt-group check, which is a different calculation (Mott ch.18). NOT done here.",
    "Shaft 2 -> 13T chain sprocket:  6-TOOTH SPLINE, not a key. Major Ø24.70, minor Ø20.88, ~25 mm long "
    "(x 84.0-109.6), with a 5 mm A/F hex socket in the end for a retaining bolt. Splines are Mott section "
    "11-5, a different torque-capacity equation. NOT done here -- flagged so it does not get forgotten.",
    "Shaft 2 sprocket is OVERHUNG: its centre sits ~24 mm outboard of bearing 2, so the chain pull is a "
    "cantilever load on the shaft, not a between-bearings load. The Shafts sheet handles this separately.",
]:
    ws[f"A{r}"] = txt
    ws[f"A{r}"].font = GREY
    ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{r}:F{r}")
    ws.row_dimensions[r].height = 42
    r += 1

# --- summary table at the top-right so the whole thing reads at a glance
sr = 4
ws["H3"] = "SUMMARY"
ws["H3"].font = H2
ws["H3"].fill = F_TITLE
for col, txt in zip("HIJKLMN", ["Station", "T (Nm)", "D", "W", "H", "Lmin", "Lhub"]):
    c = ws[f"{col}{sr}"]
    c.value, c.font, c.fill = txt, HDR, F_HDR
ws["O4"] = "FOS"; ws["O4"].font = HDR; ws["O4"].fill = F_HDR
ws["P4"] = "Fits?"; ws["P4"].font = HDR; ws["P4"].fill = F_HDR
for i, (name, rT, rD, rW, rH, rL, rLs, rLb, rLr, rN, rFit) in enumerate(key_rows):
    rr = sr + 1 + i
    ws[f"H{rr}"] = name
    ws[f"I{rr}"] = f"=F{rT}/1000"
    ws[f"J{rr}"] = f"=F{rD}"
    ws[f"K{rr}"] = f"=F{rW}"
    ws[f"L{rr}"] = f"=F{rH}"
    ws[f"M{rr}"] = f"=F{rLr}"
    ws[f"N{rr}"] = f"=F{rL}"
    ws[f"O{rr}"] = f"=F{rN}"
    ws[f"P{rr}"] = f"=F{rFit}"
    for col in "IJKLMNO":
        ws[f"{col}{rr}"].number_format = "0.0"
    for col in "HIJKLMNOP":
        ws[f"{col}{rr}"].border = BOX
widths(ws, {"H": 38, "I": 9, "J": 7, "K": 7, "L": 7, "M": 9, "N": 9, "O": 8, "P": 20})

# ================================================================== GEOMETRY
ws = wb.create_sheet("Geometry (from STEP)")
widths(ws, {"A": 34, "B": 15, "C": 15, "D": 15, "E": 15, "F": 66})
ws["A1"] = "MEASURED GEOMETRY  --  read directly out of the SolidWorks STEP files"
ws["A1"].font = H1
ws["A1"].fill = F_TITLE
ws["A2"] = ("Nothing on this sheet is assumed. Every number was extracted from the B-rep: cylindrical "
            "surfaces give the diameters and journal lengths, planes parallel to the axis give the "
            "keyseat width and depth, planes normal to it give the shoulders.")
ws["A2"].font = GREY
ws["A3"] = ("Source: motor_to_gear_shaft.STEP (authored in INCH, converted x25.4) and "
            "shaft_to_sprocket.STEP (authored in MM). Axis = X in both.")
ws["A3"].font = GREY

r = 5
def gline(r, a, b, c="", note="", font=None, fill=None):
    ws[f"A{r}"] = a
    ws[f"B{r}"] = b
    ws[f"C{r}"] = c
    ws[f"F{r}"] = note
    ws[f"F{r}"].font = GREY
    ws[f"F{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    if font:
        ws[f"A{r}"].font = font
    if fill:
        for col in "ABC":
            ws[f"{col}{r}"].fill = fill
    return r + 1

r = gline(r, "SHAFT 1 -- motor_to_gear_shaft.STEP", "", "", "Overall length 97.00 mm.", H2, F_TITLE)
r = gline(r, "Feature", "Dimension", "Axial x (mm)", "Note", HDR)
r = gline(r, "Motor flange", "D 94.00 x 7.00", "0.0 - 7.0", "12 x D9.00 holes on a D75.0 bolt circle. This joint is BOLTED, not keyed.")
r = gline(r, "Pilot boss", "D 56.00 x 2.30", "0.0 - 2.3", "Spigot into the motor face.")
r = gline(r, "Journal", "D 22.00 x 9.00", "13.0 - 22.0", "")
r = gline(r, "Bearing seat 1", "D 17.00 x 11.40", "22.6 - 34.0", "17 mm bore bearing. Centre at x = 28.3.")
r = gline(r, "Relief groove", "D 16.20 x 1.30", "34.0 - 35.3", "")
r = gline(r, "PINION JOURNAL", "D 15.00 x 36.70", "39.3 - 76.0", "This is where the 15T spur pinion sits.")
r = gline(r, "  KEYSEAT", "W 5.00 x L 35.00 overall", "40.0 - 75.0",
          "End-milled profile keyseat, D5 cutter. The CUTTER CENTRES run x 42.5 to 72.5 (30.0 mm of travel), so the finished slot is 30.0 + 5.0 = 35.0 mm overall -- do not mistake the centre travel for the slot length. Depth 2.93 mm (shaft radius 7.50 minus floor at 4.57); DIN 6885-1 t1 for a 5x5 key is 3.0, so standard depth. Mott Table 11-1 for a 15 mm shaft gives 5 x 5, so the size is correct.")
r = gline(r, "Relief groove", "D 13.80 x 1.30", "76.0 - 77.3", "")
r = gline(r, "Bearing seat 2", "D 12.00 x 16.40", "79.6 - 96.0", "12 mm bore bearing. Centre at x = 87.8.")
r = gline(r, "BEARING SPAN", "59.5 mm", "28.3 - 87.8", "Centre to centre.")
r = gline(r, "Pinion centre from brg 1", "29.2 mm", "57.5", "Keyseat midpoint.")
r += 1

r = gline(r, "SHAFT 2 -- shaft_to_sprocket.STEP", "", "", "Overall length 109.63 mm.", H2, F_TITLE)
r = gline(r, "Feature", "Dimension", "Axial x (mm)", "Note", HDR)
r = gline(r, "Bearing seat 1", "D 20.00 x 14.20", "1.0 - 15.2", "20 mm bore bearing. Centre at x = 8.1.")
r = gline(r, "Relief groove", "D 21.00 x 1.10", "19.3 - 20.4", "")
r = gline(r, "SPUR GEAR JOURNAL", "D 22.00 x 36.00", "20.5 - 56.5", "This is where the 30T spur gear sits.")
r = gline(r, "  KEYSEAT", "W 6.00 x L 30.00 overall", "23.75 - 53.75",
          "End-milled profile keyseat, D6 cutter. Cutter centres run x 26.75 to 50.75 (24.0 mm of travel), so the slot is 24.0 + 6.0 = 30.0 mm overall. Depth 3.46 mm (shaft radius 11.00 minus floor at 7.54); DIN 6885-1 t1 for a 6x6 key is 3.5, so standard depth. Mott Table 11-1 for a 22 mm shaft gives 6 x 6, size correct. The journal is 36.0 mm long, so the slot uses 30 of the 36 mm available.")
r = gline(r, "Shoulder collar", "D 30.00 x 3.00", "57.5 - 60.5", "")
r = gline(r, "Bearing seat 2", "D 25.00 x 22.00", "61.5 - 83.5", "25 mm bore bearing. Centre at x = 72.5. The 22 mm seat length suggests bearing plus a spacer.")
r = gline(r, "SPROCKET END -- SPLINE", "6 teeth, D24.70 / D20.88", "84.0 - 109.6",
          "6-tooth spline, ~25.6 mm long. NOT a key, so it is not covered by the Keys sheet -- splines are Mott section 11-5. Plus a 5 mm across-flats hex socket in the end face for a retaining bolt.")
r = gline(r, "BEARING SPAN", "64.4 mm", "8.1 - 72.5", "Centre to centre.")
r = gline(r, "Gear centre from brg 1", "30.65 mm", "38.75", "Keyseat midpoint.")
r = gline(r, "SPROCKET OVERHANG", "24.3 mm", "96.8", "Spline midpoint sits OUTBOARD of bearing 2. The chain pull is therefore a cantilever load, not a between-bearings load.")
r += 1
r = gline(r, "THE GEARS -- from CFR26 motor gearbox assembly.STEP", "", "", "", H2, F_TITLE)
r = gline(r, "Part", "Dimension", "Along axis", "Note", HDR)
r = gline(r, "15T pinion", "KHK MSGA2.5-15", "",
          "STOCK PART, which settles several open questions at once: module 2.5, pressure angle 20 deg "
          "(MSGA series standard), material SCM415 carburised with ground teeth. PCD 37.5 mm.")
r = gline(r, "  tip / rim / face", "D42.50 / D31.25 / 23.0 wide", "-11.5 to 11.5",
          "Tip diameter 42.50 = 2.5 x (15+2), which confirms module 2.5. Blank width 25.0 mm.")
r = gline(r, "  hub boss", "D30.00 x 10.75", "13.4 - 24.1", "Boss extending off one face.")
r = gline(r, "  BORE", "D15.000 x 36.25 LONG", "-12.1 - 24.1",
          "*** THE NUMBER THAT WAS BLOCKING EVERYTHING. *** Bore matches shaft 1's D15 journal exactly. "
          "At 36.25 mm the hub is LONGER than the 35.0 mm keyseat slot, so the HUB IS NOT THE CONSTRAINT "
          "-- the slot is. Every square-key option that fits the slot is covered by hub.")
r = gline(r, "30T spur gear", "KHK MSGA2.5-30", "", "Same series. PCD 75.0 mm.")
r = gline(r, "  tip / rim / face", "D80.00 / D68.75 / 23.0 wide", "-11.5 to 11.5",
          "Tip 80.00 = 2.5 x (30+2). Blank width 25.0 mm.")
r = gline(r, "  hub boss", "D50.00 x 10.57", "13.4 - 24.0", "")
r = gline(r, "  BORE", "D22.000 x 35.90 LONG", "-12.0 - 24.0",
          "Matches shaft 2's D22 journal. Hub 35.90 mm against a 30.0 mm slot, and the journal is 36.0 mm "
          "-- so extending the slot to 36 mm is fully usable, the hub covers it.")
r = gline(r, "Keyway widths in the gears", "6.00 (30T), 5.00 (15T)", "",
          "Measured off the gear bodies and they match the shaft keyseats exactly. Both correct per "
          "Mott Table 11-1 / DIN 6885-1 for their bore.")
r += 1
r = gline(r, "BEARINGS -- designations read from the assembly", "", "", "", H2, F_TITLE)
r = gline(r, "Shaft 1", "SKF 6203-2RSL + SKF 3201 A-2RS1TN9", "",
          "6203 measured in the assembly at OD 40.0 x 12.0 wide, bore 17 -- matches shaft 1's D17 seat. "
          "3201 A is a double-row angular contact, bore 12, matching the D12 seat.")
r = gline(r, "Shaft 2", "SKF 61804-2RS1 + SKF BS2-2205-2RS VT143", "",
          "61804 bore 20 matches the D20 seat; BS2-2205 is a spherical roller, bore 25, matching the D25 "
          "seat. All four bearing bores match the journals measured off the shaft files independently.")
r += 1
r = gline(r, "THE KEY PARTS THEMSELVES", "", "", "", H2, F_TITLE)
r = gline(r, "Part", "Section x length", "Ends", "Note", HDR)
r = gline(r, "Shaftkey_motor_to_gear.STEP", "5.00 x 5.00 x 27.50", "one R2.50, one square",
          "Straight (full-height) bearing length 25.00 mm. Sits in a 35.0 mm slot, so there is 7.5 mm of unused slot -- the cheapest available fix on this joint is simply a longer key.")
r = gline(r, "shaftkey_gear_to_sprocket.STEP", "6.00 x 6.00 x 27.00", "one R3.00, one square",
          "Straight bearing length 24.00 mm. Sits in a 30.0 mm slot, so the most this joint can reach without a bore change is about 27 mm straight.")
r = gline(r, "Cross-check vs Mott Table 11-1", "5x5 on D15, 6x6 on D22", "both correct",
          "Both keys are the standard section for their shaft, and both keyseats are cut to DIN 6885-1 depth. Nothing about the SIZING is wrong -- the only question is LENGTH.")
r += 1
r = gline(r, "STILL NEEDED", "", "", "", H2, F_BAD)
for txt in ["SHAFT material grade and heat treat. This is now the ONLY thing standing between a keys-only fix and new shafts -- see the Square Key Options sheet.",
            "Confirmation that this CFR26 gearbox carries over to CFR27 unchanged. The assembly is labelled CFR26.",
            "The assembly contains a JTF 1324 12T chain sprocket, but the CFR27 gear study assumes a 13T driver. Does not affect shafts 1 and 2 (they sit upstream of the chain) but it does change the diff-side torques.",
            "The diff input sprocket and diff output stub shafts -- no STEP supplied yet.",
            "The 6-tooth spline at the sprocket end of shaft 2 is still unchecked (Mott 11-5), as is the bolted motor flange (bolt group, Mott ch.18)."]:
    ws[f"A{r}"] = "  - " + txt
    ws[f"A{r}"].font = GREY
    ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{r}:F{r}")
    ws.row_dimensions[r].height = 28
    r += 1

# ================================================================== SHAFTS
ws = wb.create_sheet("Shafts")
widths(ws, {"A": 42, "B": 16, "C": 16, "D": 16, "E": 16, "F": 62, "G": 28})
ws["A1"] = "CFR27 SHAFT SIZING  --  Mott section 12-8, eq 12-24"
ws["A1"].font = H1
ws["A1"].fill = F_TITLE
ws["A2"] = ("BLUE = measured off the STEP files. YELLOW = still needed. GREEN = computed. "
            "Bending is a two-bearing model, plus a separate cantilever term for the overhung sprocket.")
ws["A2"].font = GREY

SH = ["Shaft 1 (motor)", "Shaft 2 (intermediate)", "Diff input (no STEP)"]
for j, nm in enumerate(SH):
    c = ws.cell(row=4, column=2 + j, value=nm)
    c.font, c.fill = H2, F_TITLE
ws["G4"] = "Notation"; ws["G4"].font = HDR; ws["G4"].fill = F_HDR

def srow(r, label, notation, per_shaft, note="", fill=None, num="0.00", fills=None):
    ws[f"A{r}"] = label
    ws.cell(row=r, column=7, value=notation).font = MONO
    ws[f"F{r}"] = note
    ws[f"F{r}"].font = GREY
    ws[f"F{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    for j, v in enumerate(per_shaft):
        cc = ws.cell(row=r, column=2 + j, value=v)
        cc.number_format = num
        cc.border = BOX
        f = (fills[j] if fills else fill)
        if f:
            cc.fill = f
    return r + 1

M3 = [F_MEAS, F_MEAS, F_CAD]

r = 5
r = srow(r, "Torque carried", "T (Nmm)",
         [f"=Keys!F{key_rows[0][1]}", f"=Keys!F{key_rows[1][1]}", f"=Keys!F{key_rows[2][1]}"],
         "Pulled from the Keys sheet so the two sheets can never disagree.", F_CALC, "0")
r_pcd = r
r = srow(r, "Pitch diameter of the driving element", "d (mm)", [37.5, 75.0, 121.4],
         "MEASURED: the gears are KHK MSGA2.5 stock parts, so module 2.5 and PCD = 2.5 x teeth. "
         "15T -> 37.5 mm (tip 42.5 confirms it), 30T -> 75.0 mm (tip 80.0 confirms it). Pressure angle "
         "20 deg is the MSGA series standard. The diff-input column is still a placeholder.",
         None, "0.00", [F_MEAS, F_MEAS, F_CAD])
r_Wt = r
r = srow(r, "Tangential load", "Wt = 2T/d (N)",
         [f"=2*B5/B{r_pcd}", f"=2*C5/C{r_pcd}", f"=2*D5/D{r_pcd}"], "", F_CALC, "0.0")
r_Wr = r
r = srow(r, "Radial load", "Wr = Wt*tan(phi) (N)",
         [f"=B{r_Wt}*TAN(RADIANS({PHI}))", f"=C{r_Wt}*TAN(RADIANS({PHI}))", f"=D{r_Wt}*TAN(RADIANS({PHI}))"],
         "Spur gears only. A CHAIN sprocket has no separating force -- the pull is tangential along the taut strand -- so for a chain station set this to 0 and take W = Wt.",
         F_CALC, "0.0")
r_W = r
r = srow(r, "Resultant transverse load", "W (N)",
         [f"=SQRT(B{r_Wt}^2+B{r_Wr}^2)", f"=SQRT(C{r_Wt}^2+C{r_Wr}^2)", f"=SQRT(D{r_Wt}^2+D{r_Wr}^2)"],
         "", F_CALC, "0.0")
r_span = r
r = srow(r, "Bearing span", "L (mm)", [59.5, 64.4, 160.0],
         "MEASURED centre to centre: shaft 1 brg centres x=28.3 and 87.8; shaft 2 x=8.1 and 72.5.",
         None, "0.0", M3)
r_a = r
r = srow(r, "Driving element position from brg 1", "a (mm)", [29.2, 30.65, 80.0],
         "MEASURED, keyseat midpoint: shaft 1 x=57.5, shaft 2 x=38.75.", None, "0.00", M3)
r_R1 = r
r = srow(r, "Bearing 1 reaction", "R1 (N)",
         [f"=B{r_W}*(B{r_span}-B{r_a})/B{r_span}", f"=C{r_W}*(C{r_span}-C{r_a})/C{r_span}",
          f"=D{r_W}*(D{r_span}-D{r_a})/D{r_span}"], "", F_CALC, "0.0")
r_Min = r
r = srow(r, "Bending moment, between bearings", "M inner (Nmm)",
         [f"=B{r_R1}*B{r_a}", f"=C{r_R1}*C{r_a}", f"=D{r_R1}*D{r_a}"],
         "Simply supported, single load between the bearings.", F_CALC, "0")
r += 1

ws[f"A{r}"] = "Overhung load (shaft 2's sprocket hangs outboard of bearing 2)"
ws[f"A{r}"].font = H2
r += 1
r_ovh = r
r = srow(r, "Overhang, brg 2 to element centre", "e (mm)", [0.0, 24.3, 0.0],
         "MEASURED on shaft 2: bearing 2 centre x=72.5, spline midpoint x=96.8. Zero on the others (nothing outboard).",
         None, "0.0", M3)
r_Wov = r
r = srow(r, "Overhung load (chain pull)", "W ovh (N)", [0.0, 11305.0, 0.0],
         "*** CHECK *** Shaft 2 carries the 13T sprocket here: chain pull = 2T/PCD = 2*300000/53.07 = 11305 N at the 150 Nm motor cap. Update if the sprocket or the design torque changes.",
         None, "0.0", M3)
r_Mov = r
r = srow(r, "Bending moment from the overhang", "M ovh (Nmm)",
         [f"=B{r_ovh}*B{r_Wov}", f"=C{r_ovh}*C{r_Wov}", f"=D{r_ovh}*D{r_Wov}"],
         "Cantilever: M at bearing 2 = W x e. A cantilever load does NOT relieve at the bearing, it peaks there.",
         F_CALC, "0")
r_M = r
r = srow(r, "GOVERNING bending moment", "M (Nmm)",
         [f"=MAX(B{r_Min},B{r_Mov})", f"=MAX(C{r_Min},C{r_Mov})", f"=MAX(D{r_Min},D{r_Mov})"],
         "MAX of the two, not the sum. They peak at different sections (one at the gear, one at bearing 2), so taking the larger is right for sizing a single diameter. If the two sections have DIFFERENT diameters, size each one on its own moment.",
         F_CALC, "0")
r += 1

ws[f"A{r}"] = "Endurance limit (Mott ch.5)"; ws[f"A{r}"].font = H2; r += 1
r_D = r
r = srow(r, "CHOSEN shaft diameter", "D (mm)",
         [f"=Keys!F{key_rows[0][2]}", f"=Keys!F{key_rows[1][2]}", f"=Keys!F{key_rows[2][2]}"],
         "From the Keys sheet, i.e. the keyed journal: 15.00 and 22.00 mm, both measured. This is the section the keyseat weakens, so it is the one to check.",
         F_CALC, "0.00")
r_snp = r
r = srow(r, "Endurance limit, unmodified", "sn' (MPa)", [f"=0.5*{SUSHAFT}"] * 3,
         "0.5*Su is the standard estimate for wrought steel. Mott Figure 5-8 gives it properly by surface finish -- read it off for the real machined finish before this goes in a report.",
         F_CALC, "0.0")
r_Cs = r
r = srow(r, "Size factor", "Cs = (D/7.62)^-0.11",
         [f"=(B{r_D}/7.62)^-0.11", f"=(C{r_D}/7.62)^-0.11", f"=(D{r_D}/7.62)^-0.11"],
         "Mott, D in mm.", F_CALC, "0.000")
r_CR = r
r = srow(r, "Reliability factor", "CR", [0.81] * 3, "0.81 = 99% reliability (Mott Table 5-1). 0.75 for 99.9%.", F_CAD, "0.00")
r_Cm = r
r = srow(r, "Material factor", "Cm", [1.0] * 3, "1.0 wrought steel.", F_CAD, "0.00")
r_sn = r
r = srow(r, "Modified endurance limit", "sn (MPa)",
         [f"=B{r_snp}*B{r_Cs}*B{r_CR}*B{r_Cm}", f"=C{r_snp}*C{r_Cs}*C{r_CR}*C{r_Cm}",
          f"=D{r_snp}*D{r_Cs}*D{r_CR}*D{r_Cm}"], "", F_CALC, "0.0")
r += 1

ws[f"A{r}"] = "Required diameter"; ws[f"A{r}"].font = H2; r += 1
r_Dreq = r
vals = []
for cl in "BCD":
    vals.append(f"=((32*{NDES}/PI())*SQRT(({KT}*{cl}{r_M}/{cl}{r_sn})^2+0.75*({cl}5/{SYSHAFT})^2))^(1/3)")
r = srow(r, "Required minimum diameter", "D req (mm)", vals,
         "Mott eq 12-24: D = [(32N/pi)*sqrt((Kt*M/sn)^2 + (3/4)*(T/sy)^2)]^(1/3). Reversed bending (fatigue, hence sn) with steady torsion (static, hence sy) -- correct for a shaft that spins under a one-way drive torque.",
         F_CALC, "0.00")
r_marg = r
r = srow(r, "Margin (chosen / required)", "ul",
         [f"=B{r_D}/B{r_Dreq}", f"=C{r_D}/C{r_Dreq}", f"=D{r_D}/D{r_Dreq}"],
         "Must be >= 1.00. N = 2.0 is already inside D req, so 1.00 means the design factor is exactly met, not that it is marginal.",
         F_CALC, "0.000")
r = srow(r, "Verdict", "",
         [f'=IF(B{r_D}>=B{r_Dreq},"OK","UNDERSIZED")', f'=IF(C{r_D}>=C{r_Dreq},"OK","UNDERSIZED")',
          f'=IF(D{r_D}>=D{r_Dreq},"OK","UNDERSIZED")'], "", None, "General")
r += 1

ws[f"A{r}"] = "Static torsion screen (keyseat included)"; ws[f"A{r}"].font = H2; r += 1
r_tau = r
r = srow(r, "Torsional shear at the journal", "tau = 16T/(pi*D^3) (MPa)",
         [f"=16*B5/(PI()*B{r_D}^3)", f"=16*C5/(PI()*C{r_D}^3)", f"=16*D5/(PI()*D{r_D}^3)"],
         "NO Kt, and that is deliberate. Stress-concentration factors are a FATIGUE quantity; "
         "under static load a ductile steel yields locally at the keyseat corner and redistributes, "
         "so Mott omits Kt for static loading of ductile materials. Kt belongs in the fatigue sizing "
         "above (eq 12-24, on the bending term), where it is applied.",
         F_CALC, "0.0")
r_taua = r
r = srow(r, "Allowable shear", "0.5*sy/N (MPa)", [f"=0.5*{SYSHAFT}/{NDES}"] * 3, "", F_CALC, "0.0")
r = srow(r, "FOS, static torsion", "ul",
         [f"=B{r_taua}/B{r_tau}", f"=C{r_taua}/C{r_tau}", f"=D{r_taua}/D{r_tau}"],
         "SANITY: the KEY should be the weak link, so this should come out ABOVE the key FOS on the Keys sheet. If a shaft FOS drops below its key FOS, the bore is too small and the shaft is protecting the key instead of the other way round.",
         F_CALC, "0.00")

r += 1
ws[f"A{r}"] = "Material sensitivity -- currently the biggest unknown"
ws[f"A{r}"].font = H2
ws[f"A{r}"].fill = F_BAD
r += 1
r = srow(r, "N vs shear yield, assumed 4140 (sy 621)", "ul",
         [f"=0.5*621/B{r_tau}", f"=0.5*621/C{r_tau}", f"=0.5*621/D{r_tau}"],
         "The design factor actually achieved against shear yield in static torsion.", F_CALC, "0.00")
r = srow(r, "N vs shear yield, 4340 OQT 1000 (sy 1090)", "ul",
         [f"=0.5*1090/B{r_tau}", f"=0.5*1090/C{r_tau}", f"=0.5*1090/D{r_tau}"],
         "The grade the CBR27 Baja shafts use. Same geometry, and the verdict flips. Until the real "
         "grade and heat treat are confirmed, no absolute statement about shaft adequacy is defensible.",
         F_CALC, "0.00")

for rr in range(4, r + 1):
    ws.row_dimensions[rr].height = 30

# ================================================================== KEY SPEC
# The sheet you actually hand to whoever makes the keys.
ws = wb.create_sheet("Key Spec (MAKE THIS)")
widths(ws, {"A": 34, "B": 8, "C": 8, "D": 8, "E": 9, "F": 8, "G": 12,
            "H": 11, "I": 11, "J": 11, "K": 12, "L": 10, "M": 12, "N": 11, "O": 40})
ws["A1"] = "KEY SPEC  --  what to cut, per joint"
ws["A1"].font = H1
ws["A1"].fill = F_TITLE
ws["A2"] = ("W and H are NOT a choice: Mott Table 11-1 / DIN 6885-1 fix them from the shaft diameter. "
            "The only things you pick are LENGTH and END FORM. Everything here is live -- change a bore "
            "on the Keys sheet and this updates.")
ws["A2"].font = GREY
ws["A4"] = ("END FORM IS THE LEVER. A radiused end does not bear over its own projection, so a "
            "square-ended key (DIN 6885 Form B) carries the same torque in a SHORTER overall length "
            "than a round-ended one. If the aim is to shorten the keys, this is where it comes from: "
            "Form B needs W less overall length than Form A for identical capacity.")
ws["A4"].font = WARN
ws.merge_cells("A4:O4")
ws.row_dimensions[4].height = 30

eq(ws, "A6", "L_governing", height=58)
eq(ws, "G6", "bearing_length", height=58)
ws.row_dimensions[6].height = 62

hdr = ["Joint", "Bore D", "Width W", "Height H", "Shaft t1", "Hub t2",
       "Min BEARING len", "Form B order", "Form AB order", "Form A order",
       "Form B -> stock", "Slot overall", "Max Form B in slot", "FOS at max Form B",
       "Verdict"]
for j, txt in enumerate(hdr):
    c = ws.cell(row=9, column=1 + j, value=txt)
    c.font, c.fill = HDR, F_HDR
    c.alignment = Alignment(wrap_text=True, vertical="bottom")
ws.row_dimensions[9].height = 30

SLOT = [35.0, 30.0, None, None]     # overall keyseat length measured off the STEP files
for i, (name, rT, rD, rW, rH, rL, rLs, rLb, rLr, rN, rFit) in enumerate(key_rows):
    rr = 10 + i
    D = f"Keys!F{rD}"
    W = f"Keys!F{rW}"
    H = f"Keys!F{rH}"
    Lb = f"Keys!F{rLr}"                       # governing minimum BEARING length
    ws[f"A{rr}"] = name
    ws[f"B{rr}"] = f"={D}"
    ws[f"C{rr}"] = f"={W}"
    ws[f"D{rr}"] = f"={H}"
    ws[f"E{rr}"] = f"=INDEX({KTBL_T1},MATCH({D}-0.0001,{KTBL_OVER},1))"
    ws[f"F{rr}"] = f"=INDEX({KTBL_T2},MATCH({D}-0.0001,{KTBL_OVER},1))"
    ws[f"G{rr}"] = f"={Lb}"
    ws[f"H{rr}"] = f"={Lb}"                   # Form B: full length bears
    ws[f"I{rr}"] = f"={Lb}+{W}/2"             # Form AB: one radiused end
    ws[f"J{rr}"] = f"={Lb}+{W}"               # Form A: both ends radiused
    ws[f"K{rr}"] = f"=IFERROR(INDEX({KTBL_PREF},MATCH(H{rr}-0.0001,{KTBL_PREF},1)+1),\"> 200\")"
    if SLOT[i] is not None:
        ws[f"L{rr}"] = SLOT[i]
        ws[f"L{rr}"].fill = F_MEAS
        # A square-ended key cannot occupy the slot's semicircular ends, so the most that
        # fits is the STRAIGHT portion = overall slot minus one cutter diameter (= W).
        ws[f"M{rr}"] = f"=L{rr}-{W}"
        ws[f"N{rr}"] = f"=MIN(M{rr}/{Lb}*{NDES}, {NDES})"
        ws[f"O{rr}"] = (f'=IF(M{rr}>={Lb},'
                        f'"OK - a Form B key of "&TEXT(K{rr},"0")&" mm fits and makes N="&TEXT({NDES},"0.0"),'
                        f'"SHORT - slot allows only "&TEXT(M{rr},"0.0")&" mm square-ended, N="&TEXT(N{rr},"0.00")'
                        f'&". Needs a bigger bore.")')
    else:
        for col, v in [("L", "no STEP"), ("M", "-"), ("N", "-"), ("O", "slot length unknown")]:
            ws[f"{col}{rr}"] = v
        ws[f"L{rr}"].fill = F_CAD
    for col in "BCDEFGHIJKLMN":
        ws[f"{col}{rr}"].number_format = "0.0"
    ws[f"N{rr}"].number_format = "0.00"
    for col in "ABCDEFGHIJKLMNO":
        ws[f"{col}{rr}"].border = BOX
        if col in "CDEF":
            ws[f"{col}{rr}"].fill = F_CALC

r = 10 + len(key_rows) + 1
notes = [
    ("Width W / Height H", "Mott Table 11-1 (= ISO/R 773 = DIN 6885-1), looked up live from the bore. "
     "This is the 'correct spec width' -- it is not open to choice, and both CFR27 keys already match it."),
    ("Shaft t1 / Hub t2", "DIN 6885-1 keyseat depths, for the drawing. Verified against the CAD: the "
     "D15 shaft measured 2.93 against t1 = 3.0, the D22 measured 3.46 against t1 = 3.5."),
    ("Min BEARING length", "Mott 11-2 / 11-4, the governing of shear and bearing. This is the length "
     "that must actually be in CONTACT -- it is not the length you order."),
    ("  ... and it is also the MINIMUM GEAR HUB LENGTH.", "The key only bears where the hub covers it. "
     "A 30 mm key in a 20 mm hub bears over 20 mm and the rest does nothing, however long the keyseat is. "
     "So this column is simultaneously the minimum key bearing length AND the minimum hub length. "
     "Longer keys therefore push toward WIDER gear hubs, never shorter ones -- the key does not set the "
     "gear's axial position (shoulders and spacers do), it only sets how much hub has to sit over it. "
     "Both CFR27 journals have room: shaft 1 has 36.7 mm between shoulders (x 39.3-76.0), shaft 2 has "
     "36.0 mm (x 20.5-56.5). Whether the GEARS are that wide is still unconfirmed."),
    ("Form B order length", "Square both ends. Full length bears, so order length = bearing length. "
     "SHORTEST option, and the one to use if the goal is to shorten the keys."),
    ("A note on key material", "Bearing is checked against the WEAKEST of key / shaft / hub (Inputs, "
     "'Governing bearing yield'). With the shaft assumed 4140 at 621 MPa, a key harder than 621 buys "
     "NOTHING on the bearing mode -- the shaft keyseat flank becomes the limit. So upgrading the key past "
     "4140 is wasted money unless the shaft goes up with it."),
    ("Form AB order length", "One radiused end (what both CFR27 keys are today). Add W/2."),
    ("Form A order length", "Radiused both ends. Add W. LONGEST option for the same capacity."),
    ("Form B -> stock", "Rounded UP to the next DIN 6885-1 preferred length, because keys are ordered "
     "off that list rather than cut to an arbitrary number."),
    ("Slot overall", "Overall keyseat length measured off the STEP file (end-mill CENTRE travel plus "
     "one cutter diameter -- do not mistake the centre travel for the slot)."),
    ("Max Form B in slot", "A square-ended key cannot occupy the slot's semicircular ends, so the most "
     "that fits is the STRAIGHT portion = slot minus one cutter diameter (= W). THIS is the real ceiling "
     "on a square-ended key, not the overall slot length."),
    ("FOS at max Form B", "The design factor you actually reach if you fill the slot with a square-ended "
     "key, capped at the target N. If this is below N, no end-form change saves the joint and the bore "
     "has to grow."),
]
for lbl, txt in notes:
    ws[f"A{r}"] = lbl
    ws[f"A{r}"].font = Font(bold=True, size=9)
    ws[f"B{r}"] = txt
    ws[f"B{r}"].font = GREY
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"B{r}:O{r}")
    ws.row_dimensions[r].height = 26
    r += 1

# --- torque sensitivity: the whole point of picking a design torque -----
r += 2
ws[f"A{r}"] = "TORQUE SENSITIVITY  --  what each joint achieves as the design torque moves"
ws[f"A{r}"].font = H2
ws[f"A{r}"].fill = F_TITLE
ws.merge_cells(f"A{r}:O{r}")
r += 1
ws[f"A{r}"] = ("Everything scales as 1/T, so these are exact, not interpolated. FOS is against the "
               "MAXIMUM square-ended (Form B) key the existing slot allows, i.e. the best the joint can "
               "do without re-machining a bigger bore. Blank where no STEP has been supplied.")
ws[f"A{r}"].font = GREY
ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(f"A{r}:O{r}")
ws.row_dimensions[r].height = 30
r += 1

TQ = [123, 150, 160, 170]
hdr2 = ["Joint", "max Form B (mm)"] + [f"Lmin @ {t} Nm" for t in TQ] + [f"FOS @ {t} Nm" for t in TQ]
for j, txt in enumerate(hdr2):
    c = ws.cell(row=r, column=1 + j, value=txt)
    c.font, c.fill = HDR, F_HDR
    c.alignment = Alignment(wrap_text=True, vertical="bottom")
ws.row_dimensions[r].height = 28
hrow = r
r += 1

for i, (name, rT, rD, rW, rH, rL, rLs, rLb, rLr, rN_, rFit) in enumerate(key_rows):
    rr = r + i
    D, W, H = f"Keys!F{rD}", f"Keys!F{rW}", f"Keys!F{rH}"
    # torque at THIS joint scales with the design torque, so ratio it off the live value
    Tj = f"Keys!F{rT}"
    ws[f"A{rr}"] = name
    maxB = f"M{10+i}" if SLOT[i] is not None else None
    ws[f"B{rr}"] = f"={maxB}" if maxB else "-"
    for k, t in enumerate(TQ):
        # Lmin at design torque t: scale the live governing length by t / current design torque
        col_L = get_column_letter(3 + k)
        ws[f"{col_L}{rr}"] = f"=Keys!F{rLr}*{t}/{INP}!$F$5"
        ws[f"{col_L}{rr}"].number_format = "0.0"
        col_F = get_column_letter(3 + len(TQ) + k)
        if maxB:
            ws[f"{col_F}{rr}"] = f"={maxB}/{col_L}{rr}*{NDES}"
            ws[f"{col_F}{rr}"].number_format = "0.00"
        else:
            ws[f"{col_F}{rr}"] = "-"
    ws[f"B{rr}"].number_format = "0.0"
    for j in range(len(hdr2)):
        ws.cell(row=rr, column=1 + j).border = BOX

r += len(key_rows) + 1
ws[f"A{r}"] = ("Read the FOS columns against the design factor on the Inputs sheet. A joint that clears it "
               "at 150 Nm but not at 170 is telling you the margin was coming from the torque assumption, "
               "not from the geometry.")
ws[f"A{r}"].font = GREY
ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(f"A{r}:O{r}")
ws.row_dimensions[r].height = 28

# ============================================== SQUARE KEY OPTIONS
ws = wb.create_sheet("Square Key Options")
widths(ws, {"A": 20, "B": 7, "C": 6, "D": 6, "E": 7, "F": 11, "G": 11, "H": 11,
            "I": 11, "J": 11, "K": 44})
ws["A1"] = "SQUARE KEY OPTIONS  --  every route into the 2.0-3.0 band"
ws["A1"].font = H1
ws["A1"].fill = F_TITLE
ws["A2"] = ("Every row is a SQUARE-ENDED key (DIN 6885 Form B) replacing the one-radiused-end keys in the "
            "STEP files. Square ends bear over their full length, so this is the most any given slot can "
            "deliver. Live off the design torque on the Inputs sheet.")
ws["A2"].font = GREY
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A2:K2")
ws.row_dimensions[2].height = 32

ws["A4"] = ("WHY ONLY THREE MATERIAL COLUMNS. For any key that is no wider than it is tall, "
            "Lbearing/Lshear = (sy_key/sy_bear) x (W/H) >= 1, so BEARING always governs and the key's own "
            "shear strength never enters. That means the only material number that matters is sy_bear, "
            "the WEAKEST of key / shaft / hub. The gears are SCM415 at ~800 MPa, so the hub never limits: "
            "it is the key and the shaft that set the column.")
ws["A4"].font = Font(bold=True, size=9, color="9C0006")
ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A4:K4")
ws.row_dimensions[4].height = 46

ws["A6"] = "GREEN = 2.0-3.0 target band.   RED = below 2.0.   AMBER = above 3.0, more than needed."
ws["A6"].font = Font(bold=True, size=10)
ws.merge_cells("A6:K6")

hdr = ["Joint", "Bore", "W", "H", "Slot", "Max sq. key L",
       "key 1045\\n(bearing 531)", "key 4140, shaft >=4140\\n(bearing 621)",
       "key 4340 + shaft 4340\\n(bearing 800)", "Hub avail.", "What it takes"]
for j, t in enumerate(hdr):
    c = ws.cell(row=8, column=1 + j, value=t.replace("\\n", "\n"))
    c.font, c.fill = HDR, F_HDR
    c.alignment = Alignment(wrap_text=True, vertical="bottom")
ws.row_dimensions[8].height = 42

# joint, torque cell, as-built bore, as-built slot, hub length, bores, slots
JOINTS = [
    ("Sh1 / 15T pinion", f"Keys!F{key_rows[0][1]}", 15, 35.0, 36.25, [15, 17, 18, 20], [35.0, 36.2]),
    ("Sh2 / 30T spur gear", f"Keys!F{key_rows[1][1]}", 22, 30.0, 35.90, [22, 25, 28, 30], [30.0, 36.0]),
]
BEAR = [531, 621, 800]

r = 9
first = r
for jname, Tcell, D0, slot0, hub, bores, slots in JOINTS:
    jstart = r
    for D in bores:
        for slot in slots:
            ws[f"A{r}"] = jname if r == jstart else ""
            ws[f"B{r}"] = D
            ws[f"C{r}"] = f"=INDEX({KTBL_W},MATCH(B{r}-0.0001,{KTBL_OVER},1))"
            ws[f"D{r}"] = f"=INDEX({KTBL_H},MATCH(B{r}-0.0001,{KTBL_OVER},1))"
            ws[f"E{r}"] = slot
            # a square end cannot occupy the semicircular ends an end mill leaves, AND the
            # key can never be longer than the gear hub that has to cover it
            ws[f"F{r}"] = f"=MIN(E{r}-C{r},{hub})"
            for m, syb in enumerate(BEAR):
                col = get_column_letter(7 + m)
                ws[f"{col}{r}"] = f"={syb}*B{r}*D{r}*F{r}/(4*{Tcell})"
                ws[f"{col}{r}"].number_format = "0.00"
            ws[f"J{r}"] = hub
            takes = []
            if D != D0:
                takes.append("NEW SHAFT")
            if slot != slot0:
                takes.append(f"lengthen slot to {slot:.1f}")
            ws[f"K{r}"] = ", ".join(takes) if takes else "KEYS ONLY - nothing else changes"
            if not takes:
                ws[f"K{r}"].font = Font(bold=True, color="006100")
            elif "NEW SHAFT" not in takes[0]:
                ws[f"K{r}"].font = Font(color="9C6500")
            for col in "BCDEFJ":
                ws[f"{col}{r}"].number_format = "0.0"
            for col in "ABCDEFGHIJK":
                ws[f"{col}{r}"].border = BOX
            if D == D0 and slot == slot0:
                ws[f"B{r}"].fill = F_MEAS
                ws[f"E{r}"].fill = F_MEAS
            ws[f"J{r}"].fill = F_MEAS
            r += 1
    r += 1
last = r - 1

for rng in [f"G{first}:I{last}"]:
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="between", formula=["2", "3"],
        fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(bold=True, color="006100")))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["2"],
        fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006")))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["3"],
        fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C6500")))

r = last + 2
for lbl, txt in [
    ("Max sq. key L", "The lesser of (slot minus one key width) and the gear hub length. A square-ended "
     "key cannot sit in the semicircular ends an end mill leaves, and it cannot bear beyond the hub."),
    ("Hub avail.", "MEASURED off the assembly: 36.25 mm on the 15T pinion, 35.90 mm on the 30T gear. Both "
     "are LONGER than their slots, so the hub never binds -- the slot does. That was the open question "
     "and it is now closed."),
    ("What it takes", "'KEYS ONLY' = existing shafts untouched. 'lengthen slot' = re-machining the shaft "
     "you already have, not new stock. 'NEW SHAFT' = a bigger journal, which cannot be added to an "
     "existing part."),
    ("THE DECIDING UNKNOWN", "Which column you are allowed to read is set by the SHAFT grade, and that is "
     "still unconfirmed. If the shafts are 4140 you are capped at the 621 column. If they are 4340 OQT "
     "1000 the 800 column opens up, and with a 4340 key BOTH joints reach the band with no new shafts at "
     "all -- shaft 2 only needs its slot lengthened from 30 to 36 mm. One material answer swings the whole "
     "decision from 'make new shafts' to 'make new keys'."),
    ("Also check the shaft itself", "A key fix does not help if the shaft fails first. At Ø15 the shaft is "
     "N = 1.21 in static torsion as 4140 versus 2.12 as 4340. Same material answer, same direction."),
]:
    ws[f"A{r}"] = lbl
    ws[f"A{r}"].font = Font(bold=True, size=9)
    ws[f"B{r}"] = txt
    ws[f"B{r}"].font = GREY
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"B{r}:K{r}")
    ws.row_dimensions[r].height = 34
    r += 1

# ============================================== SECOND SOURCE
ws = wb.create_sheet("Second Source")
widths(ws, {"A": 30, "B": 13, "C": 13, "D": 13, "E": 13, "F": 58})
ws["A1"] = "SECOND SOURCE  --  Hindhede, Machine Design Fundamentals"
ws["A1"].font = H1
ws["A1"].fill = F_TITLE
ws["A2"] = ("Everything in this workbook is built on Mott. This sheet checks it against a second, fully "
            "SI text and records what that text adds. Where they disagree it says so.")
ws["A2"].font = GREY
ws.merge_cells("A2:F2")

r = 4
def h2(r, t):
    ws[f"A{r}"] = t
    ws[f"A{r}"].font = H2
    ws[f"A{r}"].fill = F_TITLE
    ws.merge_cells(f"A{r}:F{r}")
    return r + 1

def line(r, a, b="", c="", d="", e="", note="", font=None):
    for col, v in zip("ABCDE", [a, b, c, d, e]):
        ws[f"{col}{r}"] = v
    ws[f"F{r}"] = note
    ws[f"F{r}"].font = GREY
    ws[f"F{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    if font:
        ws[f"A{r}"].font = font
    return r + 1

r = h2(r, "1. THE METHOD AGREES  (Hindhede eq 16-5 / 16-6 vs Mott 11-2 / 11-4)")
r = line(r, "Hindhede shear", "L = 2*T*fs/(tau_w*b*D)", "", "", "",
         "Identical in form to Mott 11-2. His worked example uses tau_w = 165 and sigma_w = 330 N/mm2, "
         "i.e. sigma_w = 2*tau_w -- exactly Mott's tau_d = 0.5*sy/N and sigma_d = sy/N. Two independent "
         "texts, same equations. The key sizing in this workbook is not method-specific.")
r = line(r, "Hindhede bearing", "L = 4*T*fs/(sigma_w*t*D)")
r += 1

r = h2(r, "2. WHAT IT ADDS  --  Table 15-4, fatigue stress concentration for keyseats (after Lipson & Juvinall)")
r = line(r, "Material", "Profile bending", "Profile torsion", "Sled-runner bend", "Sled-runner tors",
         "These are Kf, NOT Kt, so no notch-sensitivity correction is needed. This confirms the Kf = 2.0 "
         "used on the Inputs sheet for a profile keyseat in quenched and tempered steel.", HDR)
r = line(r, "Annealed steel", 1.6, 1.3, 1.3, 1.3)
r = line(r, "Quenched and tempered", 2.0, 1.6, 1.6, 1.6,
         "Our shafts, whether 4140 or 4340, are Q&T -> use this row.")
r = line(r, "NOTE on sled-runner", "", "", "", "",
         "A sled-runner (circular-cutter) keyseat has Kf 1.6 in bending against 2.0 for a profile keyseat. "
         "Both CFR27 keyseats are PROFILE (end-milled, confirmed by the D5 and D6 cutter radii in the "
         "STEP). Switching to sled-runner would cut the bending stress raiser by 20% for free, if the "
         "run-out at the ends can be tolerated.")
r += 1

r = h2(r, "3. WHAT IT ADDS  --  keyseats reduce TORSIONAL STIFFNESS (Hindhede 15-11, after Hopkins)")
r = line(r, "Effective diameter", "d_eff = d - keyseat depth", "", "", "",
         "For twist over the keyseat length, use the shaft diameter LESS the keyseat depth. For standard "
         "square keys (width d/4, depth d/8) the angular twist over that length is 71% GREATER than an "
         "unkeyed shaft. Nothing in this workbook checks driveline windup, and nothing needs to for "
         "strength -- but it is a real effect on a car where shaft compliance feeds the traction control.")
r += 1

r = h2(r, "4. WHERE THEY DISAGREE  --  key size vs shaft diameter (Hindhede Table 16-3 vs Mott 11-1 / DIN 6885-1)")
r = line(r, "Shaft diameter", "Hindhede b", "Mott / DIN", "Our part", "", "", HDR)
r = line(r, "12 - 15 mm", 3, "5 x 5 (12-17)", "5 x 5 on D15",
         "", "DISAGREES. Hindhede's Table 16-3 is a coarser US-textbook table; Mott Table 11-1 is "
         "ISO/R 773 = DIN 6885-1 = JIS B 1301, which is what stock keys and stock gears are actually made "
         "to. Our 5 x 5 is LARGER than Hindhede would call for, i.e. conservative, and it matches the "
         "measured hardware. Keep DIN.")
r = line(r, "15 - 20 mm", 4, "5 x 5 / 6 x 6", "")
r = line(r, "20 - 30 mm", 6, "6 x 6 (17-22)", "6 x 6 on D22", "", "AGREES on the D22 joint.")
r = line(r, "30 - 40 mm", 8, "8 x 7 (22-30)", "")
r += 1

r = h2(r, "5. IT VALIDATES THE 170 Nm DESIGN TORQUE  (Hindhede 1-11, service factors)")
r = line(r, "Hindhede 1-11", "P_design = k_s * P_required", "", "", "",
         "\"The inherent characteristics of the power source, the driven machine, and the interaction of "
         "the two create conditions that generally make the actual load greater by 10 to 300% than that "
         "obtained from the power equations.\" Service factors k_s run from 1 to about 5.")
r = line(r, "Our overload factor", "170/150 = 1.13", "", "", "",
         "So the design torque IS a service factor, applied exactly the way the book applies one, and at "
         "1.13 it sits at the very BOTTOM of the book's 10-300% range. Stacking it on the design factor N "
         "is correct practice, not double-counting -- and if anything 1.13 is modest for a drivetrain that "
         "sees wheel hop and kerb strikes.")
r += 1

r = h2(r, "6. THE UNCHECKED SPLINE  --  Hindhede 16-6, and this is still NOT resolved")
r = line(r, "Our joint", "6-tooth parallel-sided", "", "", "",
         "Hindhede Fig 16-18 is literally a six-tooth parallel-sided spline, so the geometry class is "
         "right. Measured: major D24.70, minor D20.88, length 25.6 mm -> tooth height 1.91 mm, mean "
         "radius 11.395 mm, total flank area 293 mm2.")
r = line(r, "Capacity equation", "T = p * n_eff * h * L * R_m", "", "", "",
         "p = allowable flank bearing pressure, n_eff = number of teeth actually sharing load.")
r = line(r, "Demand", "340 N.m", "", "", "", "At the 170 Nm design torque (motor x spur 2.0).")
r = line(r, "On the SAE rating basis", "6 - 23 N.m", "", "", "",
         "The SAE parallel-spline tables the book points to are built on 1000 psi (6.9 MPa) flank "
         "pressure. On that basis this spline rates in the TENS of N.m, far under the demand. But that "
         "basis is for SLIDING splines in continuous duty with misalignment, and it is known to be very "
         "conservative for a fixed connection -- so this is NOT a failure verdict.")
r = line(r, "On a strength basis", "222 - 519 N.m", "", "", "",
         "Flank bearing at sigma_d = sy/N, with 25% to 50% of the six teeth sharing load (the usual "
         "bracket for straight-sided splines, which do not self-centre and index imperfectly). The "
         "demand of 340 N.m sits INSIDE that bracket, which is exactly why it cannot be called either way.")
r = line(r, "TO RESOLVE", "SAE J498 / Machinery's Handbook pp.1016-1019", "", "", "",
         "*** STILL OPEN. *** Neither Mott nor Hindhede carries the parallel-spline tables; both defer to "
         "Machinery's Handbook. Note also that this is a stock JT sprocket interface, so it is a proven "
         "part in motorcycle service -- but that is evidence, not analysis, and the donor bike's "
         "countershaft torque is unknown.")

for rr in range(4, r):
    ws.row_dimensions[rr].height = 30

# ================================================================== FORMULAS
ws = wb.create_sheet("Formulas")
widths(ws, {"A": 3, "B": 96})
ws["A1"] = "EVERY EQUATION THIS WORKBOOK USES"
ws["A1"].font = H1
ws["A1"].fill = F_TITLE
ws["B2"] = ("Mott, Machine Elements in Mechanical Design, 6th ed. Rendered by "
            "tools/make_key_eq_images.py -- drawn ourselves rather than cropped out of the book, "
            "since the repo is public and book figures are someone else's copyright. Same equations, "
            "and every one cites its Mott number so it can be checked against the text.")
ws["B2"].font = GREY
ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 44

SECTIONS = [
    ("KEYS  --  Mott section 11-4", None),
    (None, "shear_stress"),
    (None, "tau_design"),
    (None, "L_shear"),
    (None, "bearing_stress"),
    (None, "sigma_design"),
    (None, "L_bearing"),
    ("The two combined -- what the Keys sheet actually evaluates", None),
    (None, "L_governing"),
    (None, "N_actual"),
    (None, "which_governs"),
    ("Key end form  --  DIN 6885-1", None),
    (None, "bearing_length"),
    ("SHAFTS  --  Mott section 12-8 and chapter 5", None),
    (None, "gear_loads"),
    (None, "bending"),
    (None, "endurance"),
    (None, "shaft_dia"),
    (None, "torsion"),
]
r = 4
for title, img in SECTIONS:
    if title:
        ws[f"B{r}"] = title
        ws[f"B{r}"].font = H2
        ws[f"B{r}"].fill = F_TITLE
        ws.row_dimensions[r].height = 20
        r += 2
    else:
        eq(ws, f"B{r}", img, height=60)
        ws.row_dimensions[r].height = 64
        r += 1

wb.save(OUT)
print("wrote", OUT)
print("sheets:", wb.sheetnames)
