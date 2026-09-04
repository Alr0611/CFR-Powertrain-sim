"""Builds CFR27_Sprocket_Study.xlsx -- live formulas, PASS/FAIL, custom tooth input.

Every input cell traces to the CFR24 sheets, the Emrax 208, or Mott. Nothing invented.
Centre distance and packaging clearance are NOT in any CFR24 sheet, so they are left
blank and every check that depends on them says so instead of guessing.
"""
import csv
import os

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as get_col

HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D9E2F3")
INP = PatternFill("solid", fgColor="FFF2CC")   # yellow = type here
UNK = PatternFill("solid", fgColor="FCE4D6")   # orange = measure this
CUR = PatternFill("solid", fgColor="E2EFDA")   # green = on the car now
WF = Font(color="FFFFFF", bold=True)
B = Font(bold=True)
IT = Font(italic=True, size=9, color="555555")
_t = Side(style="thin", color="B0B0B0")
BOX = Border(left=_t, right=_t, top=_t, bottom=_t)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
wb = openpyxl.Workbook()


def widths(ws, spec):
    for col, val in spec.items():
        ws.column_dimensions[col].width = val


def banner(ws, row, text, span):
    c = ws.cell(row, 1, text)
    c.font = WF
    c.fill = HDR
    c.alignment = Alignment(vertical="center", horizontal="left")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 20


# ============================ SHEET 1: CALC ============================
ws = wb.active
ws.title = "Sprocket Calc"
COLW = {"A": 15, "B": 13, "C": 11, "D": 11, "E": 14, "F": 11, "G": 11, "H": 12,
        "I": 11, "J": 11, "K": 10, "L": 11, "M": 11, "N": 10, "O": 14, "P": 19, "Q": 34}
widths(ws, COLW)
NC = 17  # columns A..Q

banner(ws, 1, "CFR27 FINAL DRIVE SPROCKET STUDY   |   chain geometry per Mott Ch.7", NC)
n = ws.cell(2, 1, "Yellow = type in it. Orange = still unknown, must be measured. Green = what is on the car "
                  "now. The top table keeps the driver SET at 13T because it sits on a 6-lobe spline and is a "
                  "bought part. The playground at the bottom lets you move BOTH sprockets.")
n.font = IT
n.alignment = WRAP
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
ws.row_dimensions[2].height = 30

# ------------------------------- inputs -------------------------------
banner(ws, 4, "FIXED INPUTS", NC)
inputs = [
    ("Chain number", 520, "", "MEASURED-CAD", "chain.STEP, roller dia 10.1600 and pitch 15.8750 both measured"),
    ("Chain pitch P", 15.875, "mm", "MEASURED-CAD", "chain.STEP, roller centres 15.8750 apart"),
    ("Chain roller dia", 10.160, "mm", "MEASURED-CAD", "chain.STEP DT-P2131_6261K244"),
    ("Chain plate half height", 7.2517, "mm", "MEASURED-CAD", "plate end arc R, how far the chain stands proud of the pitch circle"),
    ("Driver teeth N1 (SET)", 13, "T", "MEASURED-CAD", "DT-P2120.STEP, OD 73.9324 = Mott 13T exactly. 6-lobe spline, bought part."),
    ("Gearbox ratio", 2.000, "", "from-SHEET", "15:30, Gear Design!D15/D16/D17"),
    ("Motor peak torque", 150, "Nm", "OWNER", "Emrax 208"),
    ("Motor transient torque", 170, "Nm", "OWNER", "steps up to 160-170 in enduro/accel"),
    ("Motor peak power", 68, "kW", "OWNER", "Emrax 208"),
    ("Chain avg tensile", 27.134, "kN", "DATASHEET", "Mott T7-12 no.50 = 6100 lb, conservative stand-in for 520"),
    ("Service factor", 1.3, "", "DATASHEET", "Mott T7-17, electric motor + moderate shock"),
    ("Min wrap angle", 120, "deg", "DATASHEET", "Mott guidance, flag below this"),
]
r = 5
for lbl, val, unit, tag, src in inputs:
    ws.cell(r, 1, lbl).font = B
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(r, 3, val)
    c.fill = INP
    c.border = BOX
    c.alignment = CTR
    ws.cell(r, 4, unit).alignment = CTR
    ws.cell(r, 5, tag).font = IT
    ws.cell(r, 6, src).font = IT
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=NC)
    r += 1

P = "$C$6"
ROLL = "$C$7"
PH = "$C$8"
N1 = "$C$9"
GB = "$C$10"
TPK, TTR, PWR = "$C$11", "$C$12", "$C$13"
TENS, SF, WMIN = "$C$14", "$C$15", "$C$16"

r += 1
banner(ws, r, "MEASURED / STILL TO MEASURE", NC)
mr = r + 1
for lbl, val, note in [
    ("Centre distance C", 152.08,
     "MEASURED-CAD, SolidWorks Measure, 13T centre to diff sprocket centre. Confirm once on the car."),
    ("Max allowed envelope radius", None,
     "STILL UNKNOWN. Ask the chassis lead: how much radius is there around the diff sprocket before you hit "
     "structure? The Fit check below stays blank until this is filled in."),
]:
    ws.cell(mr, 1, lbl).font = B
    ws.merge_cells(start_row=mr, start_column=1, end_row=mr, end_column=2)
    c = ws.cell(mr, 3, val)
    c.fill = INP if val is not None else UNK
    c.border = BOX
    c.alignment = CTR
    ws.cell(mr, 4, "mm").alignment = CTR
    ws.cell(mr, 5, note).font = IT
    ws.cell(mr, 5).alignment = WRAP
    ws.merge_cells(start_row=mr, start_column=5, end_row=mr, end_column=NC)
    mr += 1
C = "$C$" + str(r + 1)
ENVMAX = "$C$" + str(r + 2)
ws.row_dimensions[r + 2].height = 26

HEADS = ["Driver teeth", "Driven teeth", "Total ratio", "Driver pitch dia (mm)", "Driven pitch dia (mm)",
         "Driven OD (mm)", "Driven OD radius (mm)", "Chain envelope radius (mm)", "Wrap on driver (deg)",
         "Chain length (pitches)", "Even count", "C for even (mm)", "Axle move (mm)",
         "Wrap check", "Fit check", "Offset link?", "VERDICT"]


def header_row(hr):
    for i, h in enumerate(HEADS, 1):
        c = ws.cell(hr, i, h)
        c.font = WF
        c.fill = HDR
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        c.border = BOX
    ws.row_dimensions[hr].height = 52


def calc_row(rr, drv, dvn):
    """drv/dvn are cell refs (e.g. 'A24'). Writes columns C..Q."""
    g = "OR(%s=\"\",%s=\"\")" % (drv, dvn)          # blank guard
    Dd = "(%s/SIN(PI()/%s))" % (P, drv)
    DD = "(%s/SIN(PI()/%s))" % (P, dvn)
    # quadratic for C at an even pitch count: a*C^2 + b*C + c = 0
    aq = "(2/%s)" % P
    bq = "((%s+%s)/2-K%d)" % (drv, dvn, rr)
    cq = "(((%s-%s)/(2*PI()))^2*%s)" % (dvn, drv, P)
    ws.cell(rr, 3, "=IF(%s,\"\",%s*%s/%s)" % (g, GB, dvn, drv))
    ws.cell(rr, 4, "=IF(%s,\"\",%s)" % (g, Dd))
    ws.cell(rr, 5, "=IF(%s,\"\",%s)" % (g, DD))
    ws.cell(rr, 6, "=IF(%s,\"\",%s*(0.6+1/TAN(PI()/%s)))" % (g, P, dvn))
    ws.cell(rr, 7, "=IF(%s,\"\",F%d/2)" % (g, rr))
    ws.cell(rr, 8, "=IF(%s,\"\",E%d/2+%s)" % (g, rr, PH))
    ws.cell(rr, 9, "=IF(%s,\"\",180-2*DEGREES(ASIN(MIN(1,(E%d-D%d)/(2*%s)))))" % (g, rr, rr, C))
    ws.cell(rr, 10, "=IF(%s,\"\",2*%s/%s+(%s+%s)/2+((%s-%s)/(2*PI()))^2*%s/%s)"
            % (g, C, P, drv, dvn, dvn, drv, P, C))
    ws.cell(rr, 11, "=IF(%s,\"\",2*ROUND(J%d/2,0))" % (g, rr))
    ws.cell(rr, 12, "=IF(%s,\"\",(-%s+SQRT(%s^2-4*%s*%s))/(2*%s))" % (g, bq, bq, aq, cq, aq))
    ws.cell(rr, 13, "=IF(%s,\"\",L%d-%s)" % (g, rr, C))
    ws.cell(rr, 14, "=IF(%s,\"\",IF(I%d>=%s,\"PASS\",\"FAIL\"))" % (g, rr, WMIN))
    ws.cell(rr, 15, "=IF(%s,\"\",IF(%s=\"\",\"NEED ENVELOPE\",IF(H%d<=%s,\"PASS\",\"FAIL\")))"
            % (g, ENVMAX, rr, ENVMAX))
    # odd alternative: nearest odd pitch count, and whether it is meaningfully closer
    ws.cell(rr, 16, "=IF(%s,\"\",IF(ABS(J%d-(2*ROUND((J%d-1)/2,0)+1))<ABS(J%d-K%d)-0.15,"
                    "\"offset link is closer\",\"no, use even\"))" % (g, rr, rr, rr, rr))
    ws.cell(rr, 17, "=IF(%s,\"type both tooth counts\",IF(N%d=\"FAIL\",\"FAIL - wrap too low\","
                    "IF(O%d=\"FAIL\",\"FAIL - hits structure\",IF(O%d=\"NEED ENVELOPE\","
                    "\"ratio \"&TEXT(C%d,\"0.000\")&\", need envelope limit\","
                    "\"PASS - ratio \"&TEXT(C%d,\"0.000\")&\", move axle \"&TEXT(M%d,\"0.00\")&\" mm\"))))"
            % (g, rr, rr, rr, rr, rr, rr))
    for cc in range(1, NC + 1):
        ws.cell(rr, cc).border = BOX
        ws.cell(rr, cc).alignment = CTR
    ws.cell(rr, 3).number_format = "0.0000"
    for cc in (4, 5, 6, 7, 8, 9, 10, 12, 13):
        ws.cell(rr, cc).number_format = "0.00"


def paint(first_row, last_row):
    rng = "N%d:Q%d" % (first_row, last_row)
    for txt, bg, fg in [("FAIL", "FFC7CE", "9C0006"), ("offset link", "FFEB9C", "9C6500"),
                        ("NEED", "FCE4D6", "974706"), ("need envelope", "FCE4D6", "974706"),
                        ("PASS", "C6EFCE", "006100")]:
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=['ISNUMBER(SEARCH("%s",N%d))' % (txt, first_row)],
            fill=PatternFill("solid", fgColor=bg), font=Font(color=fg, bold=True)))


# ----------------------- table 1: driver SET at 13T -----------------------
tr = mr + 1
banner(ws, tr, "TABLE 1   |   DRIVER SET AT 13T. Sweep the driven sprocket only. This is the real buildable set.", NC)
hr = tr + 1
header_row(hr)
first = hr + 1
teeth = list(range(26, 35))
for i, N in enumerate(teeth):
    rr = first + i
    ws.cell(rr, 1, "=%s" % N1).number_format = "0"
    ws.cell(rr, 2, N)
    calc_row(rr, "A%d" % rr, "B%d" % rr)
    ws.cell(rr, 1).fill = SUB
    if N == 30:
        for cc in range(1, NC + 1):
            ws.cell(rr, cc).fill = CUR
last1 = first + len(teeth) - 1
paint(first, last1)

# --------------------------- table 2: playground ---------------------------
pr = last1 + 2
banner(ws, pr, "TABLE 2   |   PLAYGROUND. Type ANY driver and driven in the yellow cells. Both are free here.", NC)
n = ws.cell(pr + 1, 1, "Every column recalculates. Read the VERDICT column on the right. Reminder before you get "
                       "attached to a non-13T driver: it has to be a bought sprocket with the exact 6-lobe "
                       "21.0 / 25.0 / 5.0 spline bore, so check the bore before you check the ratio.")
n.font = IT
n.alignment = WRAP
ws.merge_cells(start_row=pr + 1, start_column=1, end_row=pr + 1, end_column=NC)
ws.row_dimensions[pr + 1].height = 28
hr2 = pr + 2
header_row(hr2)
pfirst = hr2 + 1
seed = [(13, 28), (13, 31), (12, 29), (14, 32), (13, 34), (15, 30), (None, None), (None, None)]
for i, (dv, dn) in enumerate(seed):
    rr = pfirst + i
    ws.cell(rr, 1, dv).fill = INP
    ws.cell(rr, 2, dn).fill = INP
    calc_row(rr, "A%d" % rr, "B%d" % rr)
    ws.cell(rr, 1).fill = INP
    ws.cell(rr, 2).fill = INP
plast = pfirst + len(seed) - 1
paint(pfirst, plast)

fn = plast + 1
c = ws.cell(fn, 1, "Chain length must be an EVEN whole number of pitches. Column K is the nearest even count, L "
                   "is the centre distance that lands exactly on it, and M is how far the diff has to move from "
                   "the 152.08 it sits at today. If M is bigger than your tensioner travel, either take the next "
                   "even count or use the offset link the P column suggests. Offset links are the weakest part "
                   "of a chain, so prefer moving the axle when you can.")
c.font = IT
c.alignment = WRAP
ws.merge_cells(start_row=fn, start_column=1, end_row=fn + 2, end_column=NC)

# ---------------------------- chain strength ----------------------------
sr = fn + 4
banner(ws, sr, "CHAIN STRENGTH   |   tension is set by the DRIVER, so it only moves if you change the driver", NC)
for j, h in enumerate(["Case", "Motor T (Nm)", "T at driver (Nm)", "Tension (kN)", "Margin vs tensile", "Check"], 1):
    c = ws.cell(sr + 1, j, h)
    c.font = WF
    c.fill = HDR
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    c.border = BOX
ws.row_dimensions[sr + 1].height = 32
for k, (lbl, src) in enumerate([("Nominal peak", TPK), ("Transient (enduro/accel)", TTR)]):
    rr = sr + 2 + k
    ws.cell(rr, 1, lbl)
    ws.cell(rr, 2, "=%s" % src)
    ws.cell(rr, 3, "=B%d*%s" % (rr, GB))
    ws.cell(rr, 4, "=2*C%d/((%s/SIN(PI()/%s))/1000)/1000" % (rr, P, N1))
    ws.cell(rr, 5, "=%s/D%d" % (TENS, rr))
    ws.cell(rr, 6, '=IF(E%d>=2,"PASS","REVIEW")' % rr)
    for cc in range(1, 7):
        ws.cell(rr, cc).border = BOX
        ws.cell(rr, cc).alignment = CTR
    for cc in (3, 4, 5):
        ws.cell(rr, cc).number_format = "0.00"
    ws.cell(rr, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.conditional_formatting.add("F%d:F%d" % (sr + 2, sr + 3), FormulaRule(
    formula=['F%d="PASS"' % (sr + 2)], fill=PatternFill("solid", fgColor="C6EFCE"),
    font=Font(color="006100", bold=True)))

dp = sr + 5
ws.cell(dp, 1, "Design power (SF x kW)").font = B
ws.merge_cells(start_row=dp, start_column=1, end_row=dp, end_column=2)
c = ws.cell(dp, 3, "=%s*%s" % (SF, PWR))
c.number_format = "0.0"
c.border = BOX
c.alignment = CTR
ws.cell(dp, 4, "kW").alignment = CTR
note = ws.cell(dp, 5, "Mott has NO no.50 power rating table (only 40 / 60 / 80). Bracketing 13T at 3000 rpm gives "
                      "~3.3 hp against a design power near 119 hp, so this chain does NOT pass Mott's power "
                      "rating. It is justified by tensile margin and short FSAE duty life, not by the table. "
                      "Mott's ratings assume 15 000 h of industrial duty. See sprocket_configs.md section 4.3.")
note.font = IT
note.alignment = WRAP
ws.merge_cells(start_row=dp, start_column=5, end_row=dp + 3, end_column=NC)

wr = dp + 5
ws.cell(wr, 1, "Chordal ripple at driver").font = B
ws.merge_cells(start_row=wr, start_column=1, end_row=wr, end_column=2)
c = ws.cell(wr, 3, "=100*(1-COS(PI()/%s))" % N1)
c.number_format = "0.0"
c.border = BOX
c.alignment = CTR
ws.cell(wr, 4, "%").alignment = CTR
n2 = ws.cell(wr, 5, "Mott flags chordal action below ~17 teeth. 13T is under that line. Kept deliberately: going "
                    "smaller makes it worse, going bigger gives back ratio you then have to buy on the driven "
                    "anyway, and the driver is a splined bought part. Inherited compromise, not a design win.")
n2.font = IT
n2.alignment = WRAP
ws.merge_cells(start_row=wr, start_column=5, end_row=wr + 2, end_column=NC)

ws.freeze_panes = "C%d" % first

# ========================== SHEET 2: SIM MAP ==========================
ws2 = wb.create_sheet("Sim Cross-Map")
widths(ws2, {"A": 14, "B": 13, "C": 13, "D": 13, "E": 15, "F": 22, "G": 46})
banner(ws2, 1, "SIM CROSS-MAP   |   each buildable ratio against the sweep we already ran", 7)
n = ws2.cell(2, 1, "HEADS UP: the sprocket ratios do NOT line up with the CSV rows the way the handoff assumed. "
                   "gear_ratio_results.csv steps 0.1, accel_results.csv steps 0.05. Only 26T (4.0000) is an exact row "
                   "in both. 34T (5.2308) is past the end of both sweeps (max 5.20) so it has no data at all. "
                   "Everything else is linear interpolation between the two bracketing rows.")
n.font = IT
n.alignment = WRAP
ws2.merge_cells("A2:G2")
ws2.row_dimensions[2].height = 46

for j, h in enumerate(["Driven teeth", "Total ratio", "0-75 m (s)", "Trap (kph)",
                       "Endurance final SOC (%)", "Sim provenance", "Note"], 1):
    c = ws2.cell(4, j, h)
    c.font = WF
    c.fill = HDR
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    c.border = BOX
ws2.row_dimensions[4].height = 34


def load(path, cols):
    out = []
    with open(path) as f:
        for d in csv.DictReader(f):
            row = {}
            for k in cols:
                try:
                    row[k] = float(d[k])
                except ValueError:
                    row[k] = None
            out.append(row)
    return sorted(out, key=lambda x: x["ratio"])


def interp(rows, col, x):
    xs = [r["ratio"] for r in rows]
    ys = [r[col] for r in rows]
    if x < xs[0] - 1e-9 or x > xs[-1] + 1e-9:
        return None, "OUT OF SWEEP"
    for i in range(len(xs) - 1):
        if xs[i] - 1e-9 <= x <= xs[i + 1] + 1e-9:
            if abs(xs[i] - x) < 1e-9:
                return ys[i], "exact"
            if abs(xs[i + 1] - x) < 1e-9:
                return ys[i + 1], "exact"
            t = (x - xs[i]) / (xs[i + 1] - xs[i])
            if ys[i] is None or ys[i + 1] is None:
                return None, "no data"
            return ys[i] + t * (ys[i + 1] - ys[i]), "interp %g-%g" % (xs[i], xs[i + 1])
    return None, "OUT OF SWEEP"


out = os.path.join(ROOT, "output")
gr = load(os.path.join(out, "gear_ratio_results.csv"), ["ratio", "SOC98"])
ac = load(os.path.join(out, "accel_results.csv"), ["ratio", "t0_75m", "trap_kph"])

rr = 5
for N in teeth:
    ratio = 2.0 * N / 13.0
    soc, sp = interp(gr, "SOC98", ratio)
    t75, tp = interp(ac, "t0_75m", ratio)
    trap, _ = interp(ac, "trap_kph", ratio)
    ws2.cell(rr, 1, N)
    ws2.cell(rr, 2, ratio).number_format = "0.0000"
    ws2.cell(rr, 3, t75 if t75 is not None else "no data").number_format = "0.000"
    ws2.cell(rr, 4, trap if trap is not None else "no data").number_format = "0.0"
    ws2.cell(rr, 5, soc if soc is not None else "no data").number_format = "0.00"
    ws2.cell(rr, 6, "exact in both CSVs" if tp == "exact" and sp == "exact" else
             ("OUT OF SWEEP" if t75 is None else "interpolated"))
    note = ""
    if N == 30:
        note = "CURRENT, on the car"
    elif N == 26:
        note = "only exact row in both CSVs, best SOC, slowest accel"
    elif N == 32:
        note = "quickest 0-75 m of the whole set"
    elif N == 34:
        note = "past the 5.20 sweep limit, rerun the sweep to 5.25 before trusting anything here"
    elif N % 2:
        note = "odd count, needs an offset link"
    ws2.cell(rr, 7, note).alignment = WRAP
    for cc in range(1, 8):
        ws2.cell(rr, cc).border = BOX
        if cc < 7:
            ws2.cell(rr, cc).alignment = CTR
    if N == 30:
        for cc in range(1, 8):
            ws2.cell(rr, cc).fill = CUR
    if N == 34:
        for cc in range(1, 8):
            ws2.cell(rr, cc).fill = UNK
    rr += 1

ws2.cell(rr + 1, 1, "SOC98 = endurance final state of charge from a 98% start, per gear_ratio_optimization.m:204. "
                    "Higher is better. 0-75 m lower is better. t0_100kph is NaN above ratio 4.5 in the CSV because "
                    "the car never reaches 100 kph, which is why 0-75 m is the accel column that matters.").font = IT
ws2.cell(rr + 1, 1).alignment = WRAP
ws2.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 3, end_column=7)

# ======================== SHEET 3: SPLINE / DRIVER ========================
ws4 = wb.create_sheet("Spline + Driver")
widths(ws4, {"A": 26, "B": 14, "C": 14, "D": 14, "E": 22, "F": 16, "G": 34})
banner(ws4, 1, "SPLINED INTERFACE   |   measured off CFR26 motor gearbox assembly.STEP", 7)
n = ws4.cell(2, 1, "Parsed straight out of the BREP solid, not eyeballed. The mating male spline is on DT-P2127, "
                   "the gearbox output shaft, which carries the same 21.0 / 25.0 cylinders, so the pair is "
                   "confirmed from both sides.")
n.font = IT
n.alignment = WRAP
ws4.merge_cells("A2:G2")
ws4.row_dimensions[2].height = 30

banner(ws4, 4, "THE CAD NAME SAYS 12T. THE GEOMETRY SAYS 13T. FOUR FEATURES AGREE.", 7)
for j, h in enumerate(["Evidence", "Measured", "12T would be", "13T would be", "Verdict"], 1):
    c = ws4.cell(5, j, h)
    c.font = WF
    c.fill = HDR
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    c.border = BOX
ev = [
    ("Outside diameter (mm)", "73.9324", "68.7713", "73.9324", "13T"),
    ("Coaxial tip faces", "13", "12", "13", "13T"),
    ("Roller seat pockets", "13", "12", "13", "13T"),
    ("Tooth flank planes", "26", "24", "26", "13T"),
]
for k, row in enumerate(ev):
    for j, v in enumerate(row, 1):
        c = ws4.cell(6 + k, j, v)
        c.border = BOX
        c.alignment = CTR
        if j == 1:
            c.alignment = Alignment(horizontal="left", vertical="center")
        if j == 5:
            c.fill = PatternFill("solid", fgColor="FFC7CE")
            c.font = Font(color="9C0006", bold=True)

banner(ws4, 11, "SPLINE DIMENSIONS   |   all MEASURED-CAD", 7)
sp = [
    ("Spline form", "6 straight-sided lobes at 60 deg"),
    ("Minor diameter", "21.000 mm"),
    ("Major diameter", "25.000 mm"),
    ("Slot / tooth width", "5.000 mm"),
    ("Sprocket plate thickness", "5.850 mm"),
    ("Retention holes", "2 x dia 6.000 mm on a 37.000 mm span"),
    ("Mating shaft part", "DT-P2127, gearbox output shaft"),
]
for k, (a, b) in enumerate(sp):
    c = ws4.cell(12 + k, 1, a)
    c.font = B
    c.border = BOX
    c2 = ws4.cell(12 + k, 2, b)
    c2.border = BOX
    ws4.merge_cells(start_row=12 + k, start_column=2, end_row=12 + k, end_column=4)

banner(ws4, 20, "DOES THE SPLINE LIMIT THE SWEEP?   NO.", 7)
n = ws4.cell(21, 1, "The DRIVEN sprocket does not touch this spline. It mounts to the differential, which is not "
                    "in this STEP at all (no part in the assembly comes near the 80.3 mm radius a 30T needs). "
                    "Sweeping the driven from 26T to 34T never changes the shaft, the spline, or the driver. "
                    "All nine configs reuse the existing splined 13T exactly as it sits.")
n.font = IT
n.alignment = WRAP
ws4.merge_cells("A21:G23")

banner(ws4, 25, "IF YOU EVER CHANGE THE DRIVER   |   the bore is not what limits you", 7)
for j, h in enumerate(["Driver teeth", "Pitch dia (mm)", "Root dia (mm)", "Wall over spline (mm)",
                       "Total ratio with 30T", "Bore OK?", "What actually limits it"], 1):
    c = ws4.cell(26, j, h)
    c.font = WF
    c.fill = HDR
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    c.border = BOX
ws4.row_dimensions[26].height = 34
for k, N in enumerate(range(10, 17)):
    rr = 27 + k
    ws4.cell(rr, 1, N)
    ws4.cell(rr, 2, "=%s/SIN(PI()/A%d)" % (P, rr)).number_format = "0.00"
    ws4.cell(rr, 3, "=B%d-10.16" % rr).number_format = "0.00"
    ws4.cell(rr, 4, "=(C%d-25)/2" % rr).number_format = "0.00"
    ws4.cell(rr, 5, "=%s*30/A%d" % (GB, rr)).number_format = "0.000"
    ws4.cell(rr, 6, '=IF(D%d>=3,"OK","TOO THIN")' % rr)
    lim = ("chordal action, Mott flags below ~17T" if N < 17 else "")
    if N < 13:
        lim = "chordal action gets worse fast below 13T"
    elif N == 13:
        lim = "CURRENT. Nothing wins by moving."
    else:
        lim = "kills reduction, buy it back on the driven instead"
    ws4.cell(rr, 7, lim).alignment = WRAP
    for cc in range(1, 8):
        ws4.cell(rr, cc).border = BOX
        if cc < 7:
            ws4.cell(rr, cc).alignment = CTR
    if N == 13:
        for cc in range(1, 8):
            ws4.cell(rr, cc).fill = CUR
ws4.conditional_formatting.add("F27:F33", FormulaRule(
    formula=['F27="OK"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)))

n = ws4.cell(35, 1, "Even a 10T leaves 8.1 mm of wall over the spline, so the bore is nowhere near the limit. "
                    "Any replacement driver still has to have this exact 6-lobe 21.0 / 25.0 / 5.0 bore, which is "
                    "a bought-part constraint, not something to machine into a blank without a broach. Check the "
                    "bore against the numbers above before ordering. Do not trust a part number alone, the one on "
                    "this sprocket already has the wrong tooth count attached to it.")
n.font = IT
n.alignment = WRAP
ws4.merge_cells("A35:G38")

# ========================== SHEET: COMBOS GRID ==========================
ws5 = wb.create_sheet("Combos Grid")
banner(ws5, 1, "ALL COMBOS   |   total ratio = 2.000 x driven / driver", 18)
n = ws5.cell(2, 1, "Rows = driven teeth (diff sprocket). Columns = driver teeth (gearbox output). "
                   "Green = inside the 4.2-4.8 band the owner expects. Yellow = inside the full 4.00-5.20 sweep "
                   "but outside that band. Blank/grey = outside the sweep entirely. The 13T column is the one "
                   "that matters, the rest is there so nobody has to ask what if.")
n.font = IT
n.alignment = WRAP
ws5.merge_cells("A2:R2")
ws5.row_dimensions[2].height = 34
ws5.column_dimensions["A"].width = 16
drivers = list(range(11, 18))
c = ws5.cell(4, 1, "driven / driver")
c.font = WF
c.fill = HDR
c.alignment = CTR
c.border = BOX
for j, dv in enumerate(drivers, 2):
    c = ws5.cell(4, j, dv)
    c.font = WF
    c.fill = HDR
    c.alignment = CTR
    c.border = BOX
    ws5.column_dimensions[get_col(j)].width = 11
c = ws5.cell(4, len(drivers) + 3, "Driven sprocket geometry")
c.font = B
ws5.merge_cells(start_row=4, start_column=len(drivers) + 3, end_row=4, end_column=len(drivers) + 6)
for j, h in enumerate(["Pitch dia (mm)", "OD (mm)", "OD radius (mm)", "Root dia (mm)"], len(drivers) + 3):
    c = ws5.cell(5, j, h)
    c.font = WF
    c.fill = HDR
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    c.border = BOX
    ws5.column_dimensions[get_col(j)].width = 14
ws5.row_dimensions[5].height = 30
for i, N2 in enumerate(range(24, 41)):
    rr = 6 + i
    c = ws5.cell(rr, 1, N2)
    c.font = B
    c.fill = SUB
    c.border = BOX
    c.alignment = CTR
    for j, dv in enumerate(drivers, 2):
        c = ws5.cell(rr, j, "='Sprocket Calc'!$B$9*$A%d/%d" % (rr, dv))
        c.number_format = "0.000"
        c.border = BOX
        c.alignment = CTR
    gj = len(drivers) + 3
    ws5.cell(rr, gj, "=15.875/SIN(PI()/$A%d)" % rr).number_format = "0.00"
    ws5.cell(rr, gj + 1, "=15.875*(0.6+1/TAN(PI()/$A%d))" % rr).number_format = "0.00"
    ws5.cell(rr, gj + 2, "=%s%d/2" % (get_col(gj + 1), rr)).number_format = "0.00"
    ws5.cell(rr, gj + 3, "=%s%d-10.16" % (get_col(gj), rr)).number_format = "0.00"
    for j in range(gj, gj + 4):
        ws5.cell(rr, j).border = BOX
        ws5.cell(rr, j).alignment = CTR
    if N2 == 30:
        ws5.cell(rr, 1).fill = CUR
last = 5 + 17
rngc = "B6:%s%d" % (get_col(len(drivers) + 1), last)
ws5.conditional_formatting.add(rngc, FormulaRule(
    formula=["AND(B6>=4.2,B6<=4.8)"], fill=PatternFill("solid", fgColor="C6EFCE"),
    font=Font(color="006100", bold=True)))
ws5.conditional_formatting.add(rngc, FormulaRule(
    formula=["AND(B6>=4,B6<=5.2)"], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C6500")))
ws5.conditional_formatting.add(rngc, FormulaRule(
    formula=["OR(B6<4,B6>5.2)"], fill=PatternFill("solid", fgColor="F2F2F2"), font=Font(color="A6A6A6")))
ws5.freeze_panes = "B6"

n = ws5.cell(last + 2, 1, "Reminder on why only the 13T column is real: the driver sits on a 6-lobe spline "
                          "(21.0 / 25.0 / 5.0) on DT-P2127 and any replacement has to have that exact bore, which "
                          "is a bought-part constraint. The driven sprocket is made in house (cfr_sprocket) and "
                          "reuses a 45.60 bore on a 92.69 BCD with 6 x dia 6.31 holes, so any tooth count in this "
                          "grid keeps the same diff mounting. That is why the driven column is the free one.")
n.font = IT
n.alignment = WRAP
ws5.merge_cells(start_row=last + 2, start_column=1, end_row=last + 5, end_column=12)

# ====================== SHEET: RATIO DECISION (FSAE POINTS) ======================
try:
    with open(os.path.join(ROOT, "output", "gear_points_model.csv")) as _f:
        PTS = list(csv.DictReader(_f))
    with open(os.path.join(ROOT, "output", "gear_meeting_matrix.csv")) as _f:
        MTX = {int(float(r["driven_teeth"])): r for r in csv.DictReader(_f)}
except FileNotFoundError:
    PTS, MTX = [], {}

if PTS:
    ws6 = wb.create_sheet("Ratio Decision")
    widths(ws6, {"A": 9, "B": 9, "C": 11, "D": 10, "E": 10, "F": 10, "G": 11, "H": 10,
                 "I": 10, "J": 10, "K": 12, "L": 12, "M": 30})
    banner(ws6, 1, "RATIO DECISION   |   FSAE Electric points vs buildable sprocket, 13T driver fixed", 13)
    n = ws6.cell(2, 1, "Concordia 2026: 476.1 pts, 19th, ENDURANCE 25.0 out of 275. The ratio moves about 6 "
                       "points across the whole 4.2-4.8 band. Finishing endurance is worth 250. Accel and "
                       "Efficiency are modelled here; Autocross (125) and Endurance time (250) are NOT, because "
                       "no lap-time-vs-ratio model exists. Treat this as a FLOOR on what the ratio is worth.")
    n.font = IT
    n.alignment = WRAP
    ws6.merge_cells("A2:M2")
    ws6.row_dimensions[2].height = 46
    heads6 = ["Driven", "Ratio", "0-75 m (s)", "Accel pts", "Wh/lap", "Effic pts", "Accel+Eff",
              "vs 30T", "Final SOC %", "Top spd", "Exits past knee %", "Grip penalty (s)", "Note"]
    for j, h in enumerate(heads6, 1):
        c = ws6.cell(4, j, h)
        c.font = WF
        c.fill = HDR
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        c.border = BOX
    ws6.row_dimensions[4].height = 40
    for i, r in enumerate(PTS):
        rr = 5 + i
        N2 = int(float(r["driven_teeth"]))
        m = MTX.get(N2, {})
        vals = [N2, float(r["ratio"]), float(r["t75_s_mu853"]), float(r["accel_pts"]),
                float(r["wh_per_lap"]), float(r["efficiency_pts"]), float(r["accel_plus_eff"]),
                float(r["vs_30T"]), float(r["SOC98"]), float(r["top_speed_kph"]),
                float(r["exits_past_knee_pct"]), float(r["grip_penalty_s"])]
        for j, v in enumerate(vals, 1):
            c = ws6.cell(rr, j, v)
            c.border = BOX
            c.alignment = CTR
            if j == 2:
                c.number_format = "0.0000"
            elif j == 3:
                c.number_format = "0.0000"
            elif j in (12,):
                c.number_format = "+0.0000;-0.0000"
            elif j == 8:
                c.number_format = "+0.0;-0.0"
            elif j > 3:
                c.number_format = "0.0"
        note = ""
        if N2 == 30:
            note = "CURRENT. Fitted, zero risk."
        elif N2 == 31:
            note = "Best in band, +1.3 pts, axle moves 0.14 mm. Needs offset link."
        elif N2 == 28:
            note = "Costs 5.0 pts, buys 0.4 pts of SOC margin. Even chain."
        elif N2 == 32:
            note = "Peak points, but 176.5 mm envelope is NOT covered by chassis."
        elif N2 in (33, 34):
            note = "Grip-sensitive, TC eats the gain. Past the sweep."
        elif N2 in (26, 27):
            note = "Too long, gives up real accel points."
        ws6.cell(rr, 13, note).alignment = WRAP
        ws6.cell(rr, 13).border = BOX
        inband = 4.2 <= float(r["ratio"]) <= 4.8
        if N2 == 30:
            for j in range(1, 14):
                ws6.cell(rr, j).fill = CUR
        elif inband:
            for j in range(1, 14):
                ws6.cell(rr, j).fill = PatternFill("solid", fgColor="EDF3FB")
    last6 = 4 + len(PTS)
    n = ws6.cell(last6 + 2, 1, "CALIBRATION: our 2026 accel score 47.8 inverts to Tyour/Tmin = 1.2228, which with "
                               "the sim time 4.6975 s puts the field best at 3.842 s (Wisconsin scored 100). "
                               "Accel is worth about 49.9 points per second at our operating point. Field best "
                               "energy is 83.7 Wh/lap (Missouri S&T, 2025 e-meter, 28 teams); we are near 240. "
                               "Sources: FSAE_2026_MI6_prelim.pdf, FSAE Rules 2026 D.9 and D.13, "
                               "output/emeter_benchmark.csv.")
    n.font = IT
    n.alignment = WRAP
    ws6.merge_cells(start_row=last6 + 2, start_column=1, end_row=last6 + 5, end_column=13)
    n = ws6.cell(last6 + 7, 1, "RECOMMENDATION: stay at 30T. 31T is the only upgrade worth arguing for (+1.3 pts, "
                               "0.14 mm axle move) but it needs an offset link and pushes exits-past-knee from 37% "
                               "to 41%, which costs autocross points this table cannot score. The real question for "
                               "the meeting is WHY endurance scored 25/275. If it was energy, 28T is the answer and "
                               "5 points is cheap insurance. If it was mechanical or thermal, the ratio is "
                               "irrelevant and 30T stays.")
    n.font = B
    n.alignment = WRAP
    ws6.merge_cells(start_row=last6 + 7, start_column=1, end_row=last6 + 10, end_column=13)

# ====================== SHEET: OPTIMUMLAP CROSS-CHECK ======================
try:
    with open(os.path.join(ROOT, "output", "optimumlap_ratio_sweep.csv")) as _f:
        OLR = list(csv.DictReader(_f))
except FileNotFoundError:
    OLR = []

if OLR:
    ws7 = wb.create_sheet("OptimumLap Check")
    widths(ws7, {"A": 10, "B": 15, "C": 15, "D": 15, "E": 14, "F": 14, "G": 13, "H": 40})
    banner(ws7, 1, "OPTIMUMLAP CROSS-CHECK   |   Michigan Endurance 2026, halfshaft 12 deg", 8)
    n = ws7.cell(2, 1, "This is the lap-time-vs-ratio model the MATLAB study did not have. Two findings: "
                       "endurance lap time is essentially ratio-invariant (0.0382 s across the whole "
                       "4.00-5.20 range, 0.054%), and the accel result is the OPPOSITE direction to the "
                       "MATLAB sim because OptimumLap runs the full 150 Nm map while gear_meeting_matrix "
                       "runs the real 123 Nm request. Neither is a bug. See output/optimumlap_crosscheck.txt.")
    n.font = IT
    n.alignment = WRAP
    ws7.merge_cells("A2:H2")
    ws7.row_dimensions[2].height = 46
    heads7 = ["Ratio", "Endurance lap (s)", "Energy (kJ/lap)", "Energy (Wh/lap)",
              "Endur TCS %", "Accel t75 (s)", "Accel TCS %", "Note"]
    for j, h in enumerate(heads7, 1):
        c = ws7.cell(4, j, h)
        c.font = WF
        c.fill = HDR
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        c.border = BOX
    ws7.row_dimensions[4].height = 34
    for i, r in enumerate(OLR):
        rr = 5 + i
        ratio = float(r["ratio"])
        for j, key in enumerate(["ratio", "endurance_lap_s", "endurance_energy_kJ",
                                 "endurance_Wh_per_lap", "endurance_TCS_pct",
                                 "accel_t75_s", "accel_TCS_pct"], 1):
            c = ws7.cell(rr, j, float(r[key]))
            c.border = BOX
            c.alignment = CTR
            c.number_format = "0.0000" if j in (2, 6) else ("0.00" if j == 1 else "0.0")
        note = ""
        if abs(ratio - 4.80) < 0.01:
            note = "fastest endurance lap of the sweep"
        elif abs(ratio - 4.00) < 0.01:
            note = "fastest accel ON THE 150 Nm MAP, least energy"
        elif abs(ratio - 4.61) < 0.01:
            note = "CURRENT"
        elif abs(ratio - 5.20) < 0.01:
            note = "worst on both, TCS active 97% of the accel run"
        ws7.cell(rr, 8, note).alignment = WRAP
        ws7.cell(rr, 8).border = BOX
        if abs(ratio - 4.61) < 0.01:
            for j in range(1, 9):
                ws7.cell(rr, j).fill = CUR
    lr = 4 + len(OLR)
    n = ws7.cell(lr + 2, 1, "WHAT IT IS WORTH: endurance total time spread is 0.84 s over 22 laps, at about "
                            "0.52 pts/s, so ENDURANCE TIME across the entire ratio sweep is worth 0.44 points. "
                            "Scaling the same 0.054% onto a 60 s autocross lap gives 0.21 points. Together the "
                            "375 points that looked unmodelled are worth about 0.6 points. Energy rises 3.07% "
                            "from 4.00 to 5.20 (MATLAB SOC said 1.2%, same direction, OptimumLap steeper).")
    n.font = IT
    n.alignment = WRAP
    ws7.merge_cells(start_row=lr + 2, start_column=1, end_row=lr + 5, end_column=8)
    n = ws7.cell(lr + 7, 1, "THE HEADLINE: the ratio decision is downstream of the TORQUE MAP decision. "
                            "Run 123 Nm and shorter gearing wins (30T-32T). Run 150 Nm and longer gearing wins "
                            "(26T-28T). Decide the torque map first. Also: measured clean launch was 4.40 s, "
                            "OptimumLap says 4.91 and the MATLAB TC sim says 4.70 at the current ratio, so both "
                            "sims are pessimistic against the one real launch we have. Trust the RANKING inside "
                            "each torque assumption, not the absolute seconds.")
    n.font = B
    n.alignment = WRAP
    ws7.merge_cells(start_row=lr + 7, start_column=1, end_row=lr + 11, end_column=8)

# ========================== SHEET 3: README ==========================
ws3 = wb.create_sheet("README")
widths(ws3, {"A": 30, "B": 100})
banner(ws3, 1, "READ THIS FIRST", 2)
lines = [
    ("What this is", "Final-drive sprocket study for CFR27. The final drive is a FIXED 15:30 gearbox (2.000) times a "
                     "chain reduction (driven/13). Only the driven sprocket is a free choice. Total ratio = 2.000 x N/13."),
    ("How to use it", "Go to the Sprocket Calc tab. Fill in the two ORANGE cells (centre distance and radial "
                      "clearance). Every PASS/FAIL then resolves. Type any tooth count into the yellow cell in the "
                      "custom row at the bottom of the config table and it gives you the ratio and a verdict."),
    ("Why cells say UNKNOWN", "Chain centre distance and packaging clearance are not in ANY of the five CFR24 "
                              "workbooks. I searched all of them. Guessing either one would decide feasibility, which "
                              "the handoff rules forbid, so they are blank and flagged instead."),
    ("Validation gate", "The sheet's pitch diameter column is =$C$11/SIN(PI()/N), which is Mott's D = P/sin(180/N). "
                        "Reproduced the 13/30 exactly: 66.3349808182 mm and 151.8725092069 mm, ratio 4.615384615384615, "
                        "error 0 to machine precision. Gate PASSED before anything was changed."),
    ("Provenance tags", "from-SHEET = read out of a CFR24 workbook. DATASHEET = Mott table or motor data. "
                        "OWNER = told to me directly. CALC = computed from tagged inputs. UNKNOWN = not available, "
                        "not measured, not guessed."),
    ("Sheet disagreement", "Gear Design!D4 and KHK!B2 say peak torque 140 Nm. Sprocket Gearing and Forces!C7 says "
                           "150 Nm. Owner says 150 Nm with 160-170 Nm transients (Emrax 208), so this workbook uses "
                           "150 nominal and 170 worst case."),
    ("Sheet mislabel", "KHK Sheet1!B13 is labelled Roller Diameter but the formula RIGHT(520,2)/10/8 returns the 520 "
                       "roller INNER WIDTH, not the roller diameter. Does not affect this study, worth knowing."),
    ("Chain caveat", "520 motorcycle chain is not ANSI no. 50. Same 5/8 in pitch, different plate and pin spec. Mott's "
                     "no. 50 tensile (6100 lb) is used as a conservative stand-in. Get the real 520 datasheet and the "
                     "margin goes up, not down."),
    ("Gear pair flag", "The fixed 15:30 pair needs Sac 1813.9 MPa at SF=1 (Gear Design!I45). Grade 2 carburised steel "
                       "is about 1550 MPa, so contact stress is MARGINAL. Pre-existing in the CFR24 sheet, does not "
                       "move across this sweep, but somebody should own it separately."),
    ("Not edited", "params_cfr26.m and gear_ratio_optimization.m were NOT touched. params_cfr26.m stays the single "
                   "source of truth for ratios. Changing the recommended ratio is a one-line edit the owner makes."),
    ("Full write-up", "sprocket_configs.md in the repo root has the full geometry, the Mott strength work, the "
                      "combined mech+sim table, and the ranked buildable verdict."),
]
r = 3
for k, v in lines:
    c = ws3.cell(r, 1, k)
    c.font = B
    c.alignment = WRAP
    c.border = BOX
    c.fill = SUB
    v2 = ws3.cell(r, 2, v)
    v2.alignment = WRAP
    v2.border = BOX
    ws3.row_dimensions[r].height = max(30, 13 * (len(v) // 95 + 1))
    r += 1

path = os.path.join(ROOT, "CFR27_Sprocket_Study.xlsx")
wb.save(path)
print("wrote", path)
print("table1 rows %d-%d, playground rows %d-%d" % (first, last1, pfirst, plast))
