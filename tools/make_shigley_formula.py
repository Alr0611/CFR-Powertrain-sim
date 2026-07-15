"""
Turn Baja's Shigley gear workbook into a FORMULA (CFR27 EV) version.

Baja:   engine -> CVT -> spur1 -> spur2 -> rear, + transfer case/front diff/bevels (AWD)
CFR27:  EMRAX 208 -----> spur gearbox (15/30) -> chain ~2.3 -> diff  (RWD, no CVT)

'First spur' reads cells BY REFERENCE, so these keep their addresses:
    module <- Parameter!B2   teeth <- Parameter!B4/C4
    speed  <- 'Motor eff'!B7  power <- !B9   torque <- !B22

Keeps:   Parameter, Motor eff, First spur, Spur gear J
Deletes: Second spur, Rear bevel, Front bevel, Bevel gear J, Bevel gear I
"""
import glob, shutil, os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
from openpyxl.styles import Font
from xl_style import (scrub, unmerge, title_bar, widths, HDR, SMALL, SMALL_I,
                      BOLD_KEY, FILL_HDR, FILL_KEY, BOX, CTR, LEFT)

SRC = [f for f in glob.glob(r"c:\Users\Aboud\Downloads\CFR27*Shigley*.xlsx") if "FORMULA" not in f][0]
OUT = r"c:\Users\Aboud\Downloads\CFR27 Shigley Gear Calcs - FORMULA.xlsx"
shutil.copy2(SRC, OUT)
wb = openpyxl.load_workbook(OUT)

# ================= Motor eff : engine + CVT -> EMRAX 208 =================
ws = wb["Motor eff"]; unmerge(ws); scrub(ws, range(1, 40), range(1, 12))
title_bar(ws, "  EMRAX 208  ·  DESIGN POINT", "  What the gears get sized against. No engine, no CVT.", "C")

# CAREFUL: 'First spur' hard-reads B7 (speed), B9 (power), B22 (torque).
# Nothing else may live in those cells, or it silently feeds the tool garbage.
# Reduction/efficiency therefore sit in B12/B13, out of the way.
rows = [
    (4,  "Motor peak torque",   140,                      "Nm",   "0"),
    (5,  "Motor max rpm",       6000,                     "rpm",  "0"),
    (6,  "Design power",        "=B4*B5*2*PI()/60/1000",  "kW",   "0.0"),
    (7,  "Speed at pinion",     "=B5/B12",                "rpm   → First spur",  "0"),
    (9,  "Power at pinion",     "=B6*B13",                "kW    → First spur",  "0.0"),
    (12, "Reduction to pinion", 1,                        "1 = motor drives the gearbox directly (no CVT)", "0"),
    (13, "Efficiency to pinion", 1,                       "1.0 = size on raw motor torque", "0.0"),
    (22, "Torque at pinion",    "=B4*B12*B13",            "Nm    → First spur",  "0"),
]
for r, lbl, val, unit, fmt in rows:
    ws.cell(r, 1, lbl).font = HDR
    c = ws.cell(r, 2, val); c.number_format = fmt; c.border = BOX; c.alignment = CTR
    ws.cell(r, 3, unit).font = SMALL

ws["A15"] = "Conservative: peak torque AND redline at once. The motor can't — but if the gears live through it, they live."
ws["A15"].font = SMALL_I
ws["A24"] = "Baja's engine / CVT / transfer case / front diff lived here. Gone — RWD EV."
ws["A24"].font = SMALL_I
widths(ws, [("A", 22), ("B", 11), ("C", 44)])
ws.sheet_view.showGridLines = False

# ================= Parameter : one spur stage + chain, RWD =================
ws = wb["Parameter"]; unmerge(ws); scrub(ws, range(1, 32), range(1, 17))
# sub=None: row 2 is Parameter!B2 = the module, which 'First spur' reads. Don't merge over it.
title_bar(ws, "  CFR27 GEARBOX  ·  SHIGLEY CHECK   —   change teeth/module, then read 'Good?'", None, "E")

# rows 4-8 must keep these addresses (First spur + Spur gear J point at B2/B4/C4)
ws["A2"] = "Module";  c = ws["B2"]; c.value = 2.5; c.number_format = "0.0"; c.border = BOX; c.alignment = CTR
ws["C2"] = "mm"
ws["A3"] = "Ratio";   c = ws["B3"]; c.value = "=C4/B4"; c.number_format = "0.000"; c.border = BOX; c.alignment = CTR
ws["C3"] = "gearbox only"
ws["A4"] = "Teeth"
for addr, val in (("B4", 15), ("C4", 30)):
    c = ws[addr]; c.value = val; c.number_format = "0"; c.border = BOX; c.alignment = CTR
ws["D4"] = "pinion | gear"
ws["A5"] = "Face width"; c = ws["B5"]; c.value = "=10*B2"; c.number_format = "0"; c.border = BOX; c.alignment = CTR
ws["C5"] = "mm"
ws["A6"] = "Good?"
ws["B6"] = '=IF(AND(\'First spur\'!E72>1,\'First spur\'!E76>1),"Yes","No")'
ws["C6"] = '=IF(AND(\'First spur\'!F72>1,\'First spur\'!F76>1),"Yes","No")'
for addr in ("B6", "C6"):
    c = ws[addr]; c.border = BOX; c.alignment = CTR; c.font = BOLD_KEY
ws["D6"] = "pinion | gear"
for a in ("A2", "A3", "A4", "A5", "A6"):
    ws[a].font = HDR
for a in ("C2", "C3", "D4", "C5", "D6"):
    ws[a].font = SMALL

ws["A8"]  = "Chain final drive"; ws["A8"].font = HDR
c = ws["B8"]; c.value = 2.305; c.number_format = "0.000"; c.border = BOX; c.alignment = CTR
ws["C8"] = "sprockets — not checked here (a chain isn't a gear)"; ws["C8"].font = SMALL
ws["A9"] = "TOTAL RATIO"; ws["A9"].font = BOLD_KEY
c = ws["B9"]; c.value = "=B3*B8"; c.number_format = "0.000"
c.border = BOX; c.alignment = CTR; c.font = BOLD_KEY; c.fill = FILL_KEY
ws["C9"] = "4.61 today · 4.2 if we chase efficiency"; ws["C9"].font = SMALL

ws["A11"] = "⚠  'Good?' currently says No — and it's the J factor, not the gearbox."
ws["A11"].font = Font(bold=True, size=9, color="9C0006")
notes = [
    "'Spur gear J' returns 0.25 for our 15T pinion. That's FILLER — Baja's chart floors at 15 teeth,",
    "so those rows are flat, not real data. Bending stress scales straight off J, so this is a false fail.",
    "Fix: read Shigley Fig 9-10 for 15T mating 30T and type it into 'Spur gear J'!B22. CFR24 used 0.325.",
    "",
    "Changing the TOTAL ratio? Change the chain sprockets (B8) — cheap. New gears are not.",
    "Changing gearbox teeth (B4/C4)? Then this sheet earns its keep — watch 'Good?'.",
    "Lower total ratio ⇒ motor makes MORE torque for the same lap ⇒ gears see more stress either way.",
]
for i, n in enumerate(notes):
    ws.cell(12 + i, 1, n).font = SMALL

widths(ws, [("A", 16), ("B", 11), ("C", 44), ("D", 14), ("E", 8)])
ws.sheet_view.showGridLines = False

# ================= relabel 'First spur' inputs (still says engine/CVT) =================
fs = wb["First spur"]
fs["A1"] = "GIVEN INFORMATION  (CFR27 — fed by the EMRAX, not an engine)"
fs["A3"] = "Torque into pinion";  fs["D3"] = "EMRAX peak torque"
fs["A4"] = "Power into pinion";   fs["D4"] = "at the design point"
fs["A5"] = "Pinion speed";        fs["D5"] = "motor rpm (direct, no CVT)"

# ================= drop the Baja-only sheets =================
for sh in ["Second spur", "Rear bevel", "Front bevel", "Bevel gear J", "Bevel gear I"]:
    if sh in wb.sheetnames:
        del wb[sh]

# 'Spur gear J' lookups for the deleted 2nd reduction -> clear
j = wb["Spur gear J"]; unmerge(j)
for r in range(16, 23):
    for c in (4, 5, 8, 9):
        j.cell(r, c).value = None
j["D16"] = "(reduction 2 removed — Formula has ONE gear stage)"
j["D16"].font = SMALL

wb.save(OUT)
print("Saved:", os.path.basename(OUT), "| sheets:", wb.sheetnames)
