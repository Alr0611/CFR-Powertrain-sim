"""
Build OUR OWN Shigley/AGMA gear check workbook. Ours, not Baja's.

Why we stopped using Baja's book: ~70 rows of formulas nobody here wrote, wired
together by raw cell address, with their materials and mislabeled rows. You can't
defend that in a design review. Every formula in THIS one is on the sheet, readable,
and cites its Shigley equation number.

It reproduces the CFR24 Driveline Tool's STRESS MATH exactly (354.0 MPa bending;
1616 vs their 1609.7 contact, 0.4% because they chart-read I and we compute it).

THE OPEN QUESTION IS PRESSURE ANGLE, and it decides everything:
    15T at 20 deg -> J = 0.245 -> bending 470 MPa   (Shigley Fig 14-6 = Mott Fig 9-10a)
    15T at 25 deg -> J = 0.345 -> bending 333 MPa   (Mott Fig 9-10b)
CFR24 used J = 0.325, which is a slightly low read of the 25 DEG chart. And 25 deg is
physically likely for us: a 15T pinion at 20 deg gets UNDERCUT (limit 17.1 teeth vs
11.2 at 25 deg). You pick 25 deg so you CAN run a small pinion.

But CFR24 contradicts itself -- its J says 25 deg while its I = 0.108 says 20 deg.
One of those is wrong. Only the gear drawing settles it. Set PA on the Gear Check
sheet (cell B9) and this workbook reads the matching chart.

Sheets:
  Gear Check  - punch in a gearset, get stresses + a verdict, with the equations shown
  Sweep       - which gearsets hit a target ratio, and how strong they come out
  J table     - BOTH J charts (20 and 25 deg), pre-interpolated for INDEX/MATCH
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from xl_style import (title_bar, widths, HDR, SMALL, SMALL_I, BOLD_KEY,
                      FILL_HDR, FILL_KEY, BOX, CTR, LEFT)
from make_eq_images import EQS, render

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "CFR27 Gear Check.xlsx")
RED = Font(bold=True, size=9, color="9C0006")

# ---------- AGMA J, digitized off the chart. TWO of them, because PRESSURE ANGLE MATTERS ----------
#
# THIS IS THE WHOLE BALLGAME. Both Shigley Fig 14-6 and Mott Fig 9-10(a) are the 20 deg
# chart -- same AGMA 218.01 data, they agree. But Mott Fig 9-10(b) is a SECOND chart for
# 25 deg spur gears, and it reads WAY higher. At 15 teeth:
#       20 deg  ->  J ~ 0.245
#       25 deg  ->  J ~ 0.345
# CFR24 used 0.325, which is a (slightly low) read of the 25 DEG chart.
#
# And there's a physical reason 25 deg is probably right for us: a 15T pinion at 20 deg
# gets UNDERCUT. Undercut limit = 2/sin^2(phi): 17.1 teeth at 20 deg, but only 11.2 teeth
# at 25 deg. You pick 25 deg precisely SO THAT you can run a small pinion.
#
# BUT the CFR24 tool is internally inconsistent about this:
#       its J = 0.325  -> says 25 deg
#       its I = 0.108  -> says 20 deg   (I at 25 deg would be 0.128)
# One of those is wrong. Only the gear drawing settles it. Set the PA on the Gear Check
# sheet and this workbook picks the matching chart.
#
# FLAT ROWS on the 20 deg table (12/15/17) are NOT filler -- below ~18 teeth the whole
# mating-gear family merges into one line. Baja's table had that right.
MATING = [17, 25, 35, 50, 85, 170, 1000]
TEETH  = [12, 15, 17, 18, 20, 24, 30, 35, 40, 50, 60, 80, 125, 275]

# --- 20 deg, standard addendum: Shigley Fig 14-6 == Mott Fig 9-10(a) ---
JTAB_20 = np.array([
    [0.207, 0.207, 0.207, 0.207, 0.207, 0.207, 0.207],   # 12  | curves MERGED here:
    [0.245, 0.245, 0.245, 0.245, 0.245, 0.245, 0.245],   # 15  | mating gear is
    [0.293, 0.293, 0.293, 0.293, 0.293, 0.293, 0.293],   # 17  | irrelevant. Real data.
    [0.300, 0.305, 0.312, 0.317, 0.326, 0.326, 0.326],   # 18  <- family fans out
    [0.312, 0.320, 0.325, 0.330, 0.339, 0.341, 0.346],   # 20
    [0.333, 0.341, 0.348, 0.355, 0.361, 0.369, 0.372],   # 24
    [0.357, 0.365, 0.374, 0.381, 0.390, 0.390, 0.394],   # 30
    [0.370, 0.380, 0.390, 0.397, 0.405, 0.405, 0.410],   # 35
    [0.380, 0.390, 0.400, 0.410, 0.419, 0.419, 0.425],   # 40
    [0.394, 0.405, 0.415, 0.436, 0.435, 0.446, 0.457],   # 50
    [0.405, 0.416, 0.429, 0.440, 0.450, 0.462, 0.471],   # 60
    [0.417, 0.430, 0.441, 0.455, 0.464, 0.480, 0.490],   # 80
    [0.430, 0.445, 0.457, 0.470, 0.482, 0.500, 0.510],   # 125
    [0.445, 0.460, 0.470, 0.490, 0.500, 0.520, 0.530],   # 275
])

# --- 25 deg, standard addendum: Mott Fig 9-10(b) ---
# Digitized by eye off a 560-dpi crop. COARSER than the 20 deg table -- treat as +/-0.01.
# The curves stay bunched at low tooth counts here (they don't fully merge like 20 deg).
JTAB_25 = np.array([
    [0.322, 0.324, 0.326, 0.328, 0.330, 0.332, 0.333],   # 12
    [0.345, 0.347, 0.349, 0.351, 0.353, 0.354, 0.355],   # 15
    [0.355, 0.358, 0.361, 0.363, 0.365, 0.367, 0.368],   # 17
    [0.360, 0.363, 0.366, 0.369, 0.372, 0.374, 0.376],   # 18
    [0.372, 0.376, 0.380, 0.384, 0.387, 0.389, 0.391],   # 20
    [0.395, 0.400, 0.405, 0.409, 0.413, 0.416, 0.418],   # 24
    [0.425, 0.431, 0.437, 0.442, 0.447, 0.450, 0.452],   # 30
    [0.440, 0.447, 0.454, 0.460, 0.465, 0.469, 0.471],   # 35
    [0.452, 0.460, 0.467, 0.474, 0.479, 0.483, 0.486],   # 40
    [0.470, 0.479, 0.487, 0.494, 0.500, 0.505, 0.508],   # 50
    [0.483, 0.492, 0.501, 0.509, 0.516, 0.521, 0.524],   # 60
    [0.500, 0.510, 0.520, 0.528, 0.536, 0.542, 0.546],   # 80
    [0.518, 0.529, 0.540, 0.549, 0.557, 0.564, 0.568],   # 125
    [0.535, 0.547, 0.558, 0.568, 0.577, 0.584, 0.589],   # 275
])

JTABS = {20: JTAB_20, 25: JTAB_25}
GRID_T = list(range(12, 61))
GRID_M = list(range(12, 101))


def J_at(nt, nm, pa=20):
    tab = JTABS[pa]
    nt = min(max(nt, TEETH[0]), TEETH[-1])
    nm = min(max(nm, MATING[0]), MATING[-1])
    col = np.array([np.interp(nt, TEETH, tab[:, k]) for k in range(len(MATING))])
    return float(np.interp(nm, MATING, col))


def undercut_limit(pa_deg):
    """Teeth below this get undercut at standard addendum. 20deg -> 17.1, 25deg -> 11.2."""
    return 2.0 / (np.sin(np.radians(pa_deg)) ** 2)


wb = openpyxl.Workbook()

# ============================ Sheet: J table ============================
jt = wb.active
jt.title = "J table"
title_bar(jt, "  AGMA J FACTOR  -  BOTH pressure angles  (standard addendum)",
              "  20 deg = Shigley Fig 14-6 = Mott Fig 9-10(a).   25 deg = Mott Fig 9-10(b).   Gear Check picks by PA.", "H")

ROW20, ROW25 = 5, 5 + len(GRID_T) + 6     # two stacked grids
for base, pa, src in ((ROW20, 20, "20 deg  -  Shigley Fig 14-6 / Mott Fig 9-10(a)"),
                      (ROW25, 25, "25 deg  -  Mott Fig 9-10(b)")):
    c = jt.cell(base - 1, 1, src); c.font = BOLD_KEY; c.fill = FILL_KEY
    c = jt.cell(base, 1, "teeth \\ mating"); c.font = HDR; c.fill = FILL_HDR; c.border = BOX
    for j, mm in enumerate(GRID_M):
        c = jt.cell(base, 2 + j, mm); c.font = HDR; c.fill = FILL_HDR; c.border = BOX; c.number_format = "0"
    for i, t in enumerate(GRID_T):
        c = jt.cell(base + 1 + i, 1, t); c.font = HDR; c.fill = FILL_HDR; c.border = BOX; c.number_format = "0"
        for j, mm in enumerate(GRID_M):
            c = jt.cell(base + 1 + i, 2 + j, round(J_at(t, mm, pa), 4)); c.number_format = "0.000"
            if pa == 20 and t < 18:
                c.fill = FILL_KEY        # amber = merged region (real data, see notes)

NOTE0 = ROW25 + len(GRID_T) + 3
jt.cell(NOTE0, 1, "PRESSURE ANGLE IS THE WHOLE BALLGAME - read this.").font = Font(bold=True, size=9, color="9C0006")
for i, n in enumerate([
    "At 15 teeth:   20 deg -> J = 0.245     25 deg -> J = 0.345.   That is a 40% swing in bending stress.",
    "CFR24 used 0.325, which is a slightly low read of the 25 DEG chart (Mott Fig 9-10 panel b).",
    "",
    "And 25 deg is probably right for us, physically: a 15T pinion at 20 deg gets UNDERCUT.",
    "Undercut limit = 2/sin^2(phi):   20 deg -> 17.1 teeth      25 deg -> 11.2 teeth.",
    "You choose 25 deg precisely so you CAN run a small pinion. 15T at 20 deg standard shouldn't exist.",
    "",
    "BUT the CFR24 tool contradicts itself:  its J = 0.325 says 25 deg, its I = 0.108 says 20 deg",
    "(I at 25 deg would be 0.128). One of those is wrong. Only the GEAR DRAWING settles it.",
    "",
    "The 20 deg grid's flat rows (12/15/17) are real: below ~18T the mating-gear curves merge into one",
    "line. The 25 deg curves stay bunched but don't fully merge. Both are AGMA 218.01 data.",
    "",
    "CAVEAT: the 25 deg grid was digitized by eye off a 560-dpi crop of Mott Fig 9-10(b). Treat it as",
    "+/-0.01. The 20 deg grid is the better-checked of the two.",
]):
    jt.cell(NOTE0 + 1 + i, 1, n).font = SMALL
widths(jt, [("A", 13)])
jt.sheet_view.showGridLines = False
LASTCOL = get_column_letter(1 + len(GRID_M))
R20a, R20b = ROW20 + 1, ROW20 + len(GRID_T)
R25a, R25b = ROW25 + 1, ROW25 + len(GRID_T)

# ============================ Sheet: Gear Check ============================
gc = wb.create_sheet("Gear Check", 0)
title_bar(gc, "  CFR27 GEAR CHECK  -  Shigley / AGMA",
              "  Punch in a gearset -> stresses and a verdict. Every formula is on this sheet.", "E")


def block(row, label):
    c = gc.cell(row, 1, label); c.font = HDR; c.fill = FILL_HDR; c.border = BOX
    for k in (2, 3):
        gc.cell(row, k).fill = FILL_HDR; gc.cell(row, k).border = BOX


def field(row, label, val, unit, fmt="0.00"):
    gc.cell(row, 1, label).font = HDR
    c = gc.cell(row, 2, val); c.number_format = fmt; c.border = BOX; c.alignment = CTR
    gc.cell(row, 3, unit).font = SMALL
    return c


block(4, "INPUTS  -  change these")
field(5,  "Pinion teeth  (Np)", 15,   "our gearbox pinion", "0")
field(6,  "Gear teeth    (Ng)", 30,   "our gearbox gear", "0")
field(7,  "Module",             2.5,  "mm", "0.0")
field(8,  "Face width",         25,   "mm", "0")
# THE INPUT THAT DECIDES EVERYTHING. 20 or 25 only -- it picks which J chart is read.
# Default 20 because that is what the CFR24 tool DECLARES ('1.1 v & Wt'!A1), and its
# I = 0.108 agrees. Its J = 0.325 does NOT agree -- see note 4. Confirm on the drawing.
c = field(9, "Pressure angle", 20, "deg  <- 20 or 25 ONLY. THIS PICKS THE J CHART. Confirm on the gear drawing.", "0")
c.fill = FILL_KEY; c.font = BOLD_KEY
gc.cell(9, 3).font = RED
# Undercut warning: 15T at 20deg is below the 17.1 limit and shouldn't exist as standard.
gc["A12"] = "Undercut limit"
gc["B12"] = "=2/(SIN(RADIANS(B9))^2)"
gc["B12"].number_format = "0.0"; gc["B12"].border = BOX; gc["B12"].alignment = CTR
gc["A12"].font = HDR
gc["C12"] = ('=IF(B5<B12,"WARNING: "&B5&"T is BELOW the undercut limit at this PA - '
             'standard teeth would be undercut. Profile shift?","OK - pinion is above the undercut limit")')
gc["C12"].font = SMALL
field(10, "Design torque",      79.6, "Nm  <- CONTINUOUS. Gears live on this, not the once-a-lap peak.", "0.0")
# Blank by default -> auto-lookup off Fig 14-6, which gives 0.245 for our 15T.
# This used to default to CFR24's 0.325, which does not survive the actual chart.
field(11, "J override",         None, "leave BLANK - the lookup reads Fig 14-6 properly now.", "0.000")

block(13, "AGMA FACTORS")
for r, (l, v, u) in enumerate([
    ("Ko  overload",           1.25,  "shock / duty"),
    ("Kv  dynamic",            1.2,   "tooth quality + pitch line speed"),
    ("Ks  size",               1.0,   ""),
    ("Km  load distribution",  1.129, "face width / alignment"),
    ("Kb  rim thickness",      1.0,   ""),
    ("Cp  elastic coeff",      191,   "MPa, steel on steel"),
]):
    field(14 + r, l, v, u, "0.000")

# THESE TWO ARE BAJA'S, NOT OURS. The CFR24 Driveline Tool computes stresses and then
# just stops - no allowables, no FOS, no verdict. So there was nothing of ours to
# divide by, and these came from Baja's 'First spur' E8/E9: their steel, their heat
# treat, their assumptions. The VERDICT still works because it compares candidate to
# baseline and the allowable CANCELS. The ABSOLUTE FOS numbers do not.
block(21, "MATERIAL ALLOWABLES   <-  BAJA'S NUMBERS. NOT OURS. SEE NOTES.")
field(22, "Allowable bending  St", 461,    "MPa - Baja's steel. Replace when we know ours.", "0")
field(23, "Allowable contact  Sc", 1627.9, "MPa - Baja's steel. Replace when we know ours.", "0.0")
for r in (22, 23):
    gc.cell(r, 3).font = RED

block(25, "RESULTS")
gc["A26"] = "Ratio";                gc["B26"] = "=B6/B5"
gc["A27"] = "Pitch diameter  Dp";   gc["B27"] = "=B5*B7";                  gc["C27"] = "mm"
gc["A28"] = "Tangential load  Wt";  gc["B28"] = "=2*B10/(B27/1000)";       gc["C28"] = "N"
gc["A29"] = "J  (bending geom)"
# Reads the 20deg grid or the 25deg grid depending on B9. This is the line that makes
# the whole CFR24 argument evaporate or not.
gc["B29"] = ("=IF(B11<>\"\",B11,IF(B9=25,"
             "INDEX('J table'!$B${r25a}:${LC}${r25b},MATCH(MIN(MAX(B5,12),60),'J table'!$A${r25a}:$A${r25b},0),"
             "MATCH(MIN(MAX(B6,12),100),'J table'!$B${h25}:${LC}${h25},0)),"
             "INDEX('J table'!$B${r20a}:${LC}${r20b},MATCH(MIN(MAX(B5,12),60),'J table'!$A${r20a}:$A${r20b},0),"
             "MATCH(MIN(MAX(B6,12),100),'J table'!$B${h20}:${LC}${h20},0))))"
             ).format(LC=LASTCOL, r20a=R20a, r20b=R20b, h20=ROW20,
                      r25a=R25a, r25b=R25b, h25=ROW25)
gc["C29"] = '=IF(B11<>"","MANUAL OVERRIDE",IF(B9=25,"Mott Fig 9-10(b), 25 deg","Shigley Fig 14-6, 20 deg"))'
gc["A30"] = "I  (contact geom)"
gc["B30"] = "=(COS(RADIANS(B9))*SIN(RADIANS(B9))/2)*(B26/(B26+1))"
gc["A31"] = "Bending stress"
gc["B31"] = "=(B28*B14*B15*B16)/(B8*B7)*(B17*B18/B29)"; gc["C31"] = "MPa"
gc["A32"] = "Contact stress"
gc["B32"] = "=B19*SQRT((B28*B14*B15*B16*B17)/(B8*B27*B30))"; gc["C32"] = "MPa"
for r in range(26, 33):
    gc.cell(r, 1).font = HDR
    c = gc.cell(r, 2); c.border = BOX; c.alignment = CTR
    c.number_format = "0.000" if r in (26, 29, 30) else "0.0"
    gc.cell(r, 3).font = SMALL

gc["A34"] = "FOS  bending"; gc["B34"] = "=B22/B31"
gc["A35"] = "FOS  contact"; gc["B35"] = "=B23/B32"
for r in (34, 35):
    gc.cell(r, 1).font = HDR
    c = gc.cell(r, 2); c.number_format = "0.00"; c.border = BOX; c.alignment = CTR; c.font = BOLD_KEY

# ---- the bar is the car, not a textbook number ----
# Our 15/30 has run for YEARS without shedding a tooth, so a "FOS > 1.5" rule that
# fails it is telling you the rule is wrong, not the gearbox. The honest use of this
# sheet is COMPARATIVE: is a candidate at least as strong as what we already run?
# That comparison is immune to the allowable being Baja's - it cancels.
# Baseline recomputes live off the same torque/factors, so it tracks your inputs.
block(37, "COMPARE TO THE CAR WE ACTUALLY RUN  (15T / 30T, same J chart as above)")
gc["A38"] = "Baseline FOS  bending"
gc["B38"] = "=B22/((2*B10/((15*B7)/1000)*B14*B15*B16)/(B8*B7)*(B17*B18/IF(B9=25,0.345,0.245)))"
gc["A39"] = "Baseline FOS  contact"
gc["B39"] = ("=B23/(B19*SQRT((2*B10/((15*B7)/1000)*B14*B15*B16*B17)"
             "/(B8*(15*B7)*((COS(RADIANS(B9))*SIN(RADIANS(B9))/2)*(2/3)))))")
for r in (38, 39):
    gc.cell(r, 1).font = HDR
    c = gc.cell(r, 2); c.number_format = "0.00"; c.border = BOX; c.alignment = CTR
gc["C38"] = '=IF(B9=25,"15T at 25 deg -> J = 0.345","15T at 20 deg -> J = 0.245")'
gc["C38"].font = SMALL
gc["C39"] = "same torque + factors as above, so it tracks your inputs"; gc["C39"].font = SMALL

gc["A41"] = "VERDICT"; gc["A41"].font = BOLD_KEY
c = gc["B41"]
c.value = ('=IF(AND(B34>=B38,B35>=B39),"OK - at least as strong as the car",'
           '"WEAKER than the car - think hard")')
c.font = BOLD_KEY; c.fill = FILL_KEY; c.border = BOX; c.alignment = CTR
gc.merge_cells("B41:C41")

# ---- show our work: the equations, next to the numbers they produce ----
gc["E25"] = "THE MATH, SHOWING ITS WORK  ->  every equation cites its Shigley number"
gc["E25"].font = HDR
for nm, (ltx, cite) in EQS.items():
    render(nm, ltx, cite)
for nm, anchor in (("wt", "E27"), ("I", "E29"), ("bend", "E31"),
                   ("cont", "E34"), ("fos", "E37")):
    img = XLImage(os.path.join(HERE, "eq_img", nm + ".png"))
    img.width, img.height = img.width * 0.42, img.height * 0.42
    gc.add_image(img, anchor)

notes = [
    "",
    "READ THIS BEFORE YOU QUOTE A NUMBER OFF THIS SHEET",
    "",
    "1) THE ALLOWABLES (B22/B23) ARE BAJA'S, NOT OURS.",
    "   The CFR24 Driveline Tool computes stresses and stops - it has no allowables and no FOS at all.",
    "   So there was nothing of ours to divide by, and 461/1627.9 came from Baja's book: their steel,",
    "   their heat treat, their assumptions. So do NOT quote 'our gearbox runs at FOS 1.3' as a fact",
    "   about our car. It is not one yet.",
    "",
    "2) THE VERDICT IS STILL GOOD - because it is a COMPARISON.",
    "   Candidate and baseline both divide by the same allowable, so the allowable CANCELS. 'Is this",
    "   gearset at least as strong as the 15/30 we already run' is valid even with Baja's steel. That",
    "   is why the verdict asks that instead of 'does it beat 1.5'. A textbook 1.5 would fail the",
    "   gearbox that has been running for years - which means the bar is wrong, not the gearbox.",
    "",
    "3) TO MAKE THE ABSOLUTE NUMBERS REAL: find out what steel our gears are and what heat treat,",
    "   then type Sat/Sac into B22/B23. That is a question for whoever spec'd the gears, not a chart",
    "   to squint at. Shigley Tables 14-3 / 14-6, carburized & hardened (the usual for gears like ours):",
    "        Grade 1:   St =  55 ksi =  379 MPa        Sc = 180 ksi = 1241 MPa",
    "        Grade 2:   St =  65 ksi =  448 MPa        Sc = 225 ksi = 1551 MPa",
    "        Grade 3:   St =  75 ksi =  517 MPa        Sc = 275 ksi = 1896 MPa",
    "   (at 1e7 cycles, 0.99 reliability). Baja's 461/1627.9 land around Grade 2. If we are actually",
    "   Grade 1, every margin here gets worse - which is exactly why someone needs to go find out.",
    "",
    "4) THE BIG ONE - THE CFR24 TOOL CONTRADICTS ITSELF, AND WE CANNOT SETTLE IT FROM THE TOOL.",
    "   Here is every hard fact we have, and they do not add up:",
    "        it DECLARES pressure angle = 20 deg      ('1.1 v & Wt'!A1)",
    "        its I = 0.108                            -> agrees, that IS the 20 deg value",
    "        its J = 0.325                            -> DISAGREES. The 20 deg chart reads 0.245 at 15T.",
    "                                                    0.325 is roughly the 25 deg value.",
    "        its centre distance C = 56.25 mm         -> exactly (15+30)*2.5/2, i.e. STANDARD",
    "   So J is the odd one out. On the tool's own declared geometry, bending should be 470 MPa, not",
    "   the 354 MPa it reports. That is a 33% understatement.",
    "",
    "   BUT THE GEARS ARE PROBABLY FINE ANYWAY, and here is the bit the tool never mentions:",
    "   a 15T pinion at 20 deg STANDARD would be UNDERCUT (limit = 2/sin^2(phi) = 17.1 teeth; row 12",
    "   flags this live). Undercut teeth are junk. So the real gears are almost certainly PROFILE",
    "   SHIFTED -- pinion positive, gear negative, which keeps C at the standard 56.25 we see. And a",
    "   positive shift ADDS root material, pushing J up from 0.245 toward... about 0.325.",
    "",
    "   So J = 0.325 may well be right for a reason nobody wrote down. Or it may be a chart misread.",
    "   Those two stories predict the same number and we cannot tell them apart from the spreadsheet.",
    "   NOTE: this sheet models STANDARD teeth only. It has no profile shift. If our gears are shifted,",
    "   the 20 deg J lookup here is PESSIMISTIC and you should use the override.",
    "",
    "   -> TWO NUMBERS OFF THE GEAR DRAWING SETTLE IT:  (1) pressure angle   (2) profile shift x.",
    "      Everything above is inference. The drawing is the only authority.",
    "",
    "   THE WAY OUT THAT DOESN'T CARE: 21T/38T. At 21 teeth you're above the undercut limit at either",
    "   pressure angle, the chart is fanned and unambiguous, it hits 4.17 total on our existing 2.305",
    "   chain, and it comes out stronger than the 15/30 whichever chart you read. No argument to have.",
    "",
    "PROOF THE STRESS MATH IS RIGHT: force J = 0.325 into B11, set PA = 20, and you get 354.0 / 1616.0",
    "MPa against the CFR24 Driveline Tool's 354.0 / 1609.7. Bending exact; contact 0.4% off because",
    "they read I = 0.108 off a chart and this computes 0.1071. So the FORMULAS agree with CFR24 -",
    "the only open question is which J chart the gears actually live on.",
    "",
    "TORQUE: gears are sized on CONTINUOUS torque (~80 Nm), not the 140 Nm peak. Check the peak too -",
    "put 140 in B10 and watch contact stress. That is the one that pits teeth if it happens often.",
    "",
    "RATIO: changing TOTAL ratio? Chain sprockets are cheap, new gears are not. But a lower ratio means",
    "the motor makes MORE torque for the same lap, so the gears see more stress either way. Check both.",
]
for i, n in enumerate(notes):
    gc.cell(43 + i, 1, n).font = SMALL
widths(gc, [("A", 24), ("B", 12), ("C", 56)])
gc.sheet_view.showGridLines = False

# ============================ Sheet: Sweep ============================
sw = wb.create_sheet("Sweep", 1)
title_bar(sw, "  GEARSET SWEEP  -  what hits the target ratio and survives",
              "  Only 18+ tooth pinions are listed. Below that you are in the merged region of Fig 14-6 and J drops fast.", "H")
sw["A4"] = "Target TOTAL ratio"; sw["A4"].font = HDR
c = sw["B4"]; c.value = 4.20; c.number_format = "0.00"
c.fill = FILL_KEY; c.font = BOLD_KEY; c.border = BOX; c.alignment = CTR
sw["A5"] = "Chain ratio"; sw["A5"].font = HDR
c = sw["B5"]; c.value = 2.305; c.number_format = "0.000"; c.border = BOX; c.alignment = CTR
sw["C5"] = "sprockets - the cheap way to change total ratio"; sw["C5"].font = SMALL
sw["A6"] = "Gearbox must do"; sw["A6"].font = HDR
c = sw["B6"]; c.value = "=B4/B5"; c.number_format = "0.000"; c.border = BOX; c.alignment = CTR

for j, h in enumerate(["Np", "Ng", "Gearbox", "TOTAL", "J", "Bending\n(MPa)", "FOS\nbend", "Verdict"]):
    c = sw.cell(8, 1 + j, h); c.font = HDR; c.fill = FILL_HDR; c.border = BOX; c.alignment = CTR
sw.row_dimensions[8].height = 28

SWEEP_PA = 25   # match the Gear Check default; reseed if you change PA there
r = 9
for Np in range(18, 25):
    for Ng in range(28, 61):
        if abs(Ng / Np - (4.20 / 2.305)) > 0.03:      # seeded at the default target
            continue
        sw.cell(r, 1, Np).number_format = "0"
        sw.cell(r, 2, Ng).number_format = "0"
        sw.cell(r, 3, "=B%d/A%d" % (r, r)).number_format = "0.000"
        sw.cell(r, 4, "=C%d*$B$5" % r).number_format = "0.000"
        sw.cell(r, 5, round(J_at(Np, Ng, SWEEP_PA), 4)).number_format = "0.000"
        sw.cell(r, 6, "=(2*'Gear Check'!$B$10/((A{0}*'Gear Check'!$B$7)/1000)*'Gear Check'!$B$14*"
                      "'Gear Check'!$B$15*'Gear Check'!$B$16)/('Gear Check'!$B$8*'Gear Check'!$B$7)*"
                      "('Gear Check'!$B$17*'Gear Check'!$B$18/E{0})".format(r)).number_format = "0"
        sw.cell(r, 7, "='Gear Check'!$B$22/F%d" % r).number_format = "0.00"
        sw.cell(r, 8, '=IF(G%d>=\'Gear Check\'!$B$38,"OK","weaker than car")' % r)
        for j in range(1, 9):
            cc = sw.cell(r, j); cc.border = BOX; cc.alignment = CTR
        r += 1

for i, n in enumerate([
    "Rows are seeded for the default target (4.20). Change B4 and re-run tools/make_gear_check_excel.py to reseed.",
    "Factors, torque and allowables all come from the Gear Check sheet - one source, no copies drifting apart.",
    "'Verdict' compares to the 15/30 baseline on Gear Check (B38), not to a textbook number. See notes there.",
]):
    sw.cell(r + 1 + i, 1, n).font = SMALL
widths(sw, [("A", 7), ("B", 7), ("C", 10), ("D", 10), ("E", 9), ("F", 11), ("G", 9), ("H", 15)])
sw.sheet_view.showGridLines = False

wb.save(OUT)
print("Saved:", os.path.basename(OUT), "| sheets:", wb.sheetnames)
for pa in (20, 25):
    print("  PA %d deg:  J(15,30) = %.3f   J(21,38) = %.3f   undercut limit = %.1f teeth"
          % (pa, J_at(15,30,pa), J_at(21,38,pa), undercut_limit(pa)))
