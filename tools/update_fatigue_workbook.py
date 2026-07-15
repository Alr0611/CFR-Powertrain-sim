"""
Rebuild the 'Load Case' sheet of Fatigue Load Cases.xlsx from REAL car data.
Always starts from the ORIGINAL backup, so re-running is safe.

  C = ACCEL (comp Jun 19)  <- the load case P2121/P2127 analyse
  D/E = endurance, reference only
  Bins now reach 160 Nm (real car hits 152; the old sheet stopped at 140)
"""
import shutil, os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.formula.translate import Translator
from xl_style import (scrub, unmerge, title_bar, widths, HDR, SMALL, BOLD_KEY,
                      FILL_HDR, FILL_KEY, BOX, CTR)

DIR = r"c:\Users\Aboud\Downloads"
SRC = os.path.join(DIR, "Fatigue Load Cases.xlsx")
BAK = os.path.join(DIR, "Fatigue Load Cases_ORIGINAL_backup.xlsx")

#         start,  end,   ACCEL,             ENDUR @4.61,        ENDUR @4.20
ROWS = [
    (0.1,  15.1, 0.0,               0.415059874162772, 0.408519787677213),
    (15.1, 30.1, 0.0,               0.170895067992697, 0.148021232278441),
    (30.1, 45.1, 0.111534795042911, 0.155943440903863, 0.143385070214343),
    (45.1, 60.1, 0.200190657769285, 0.130573032947704, 0.135725324195388),
    (60.1, 75.0, 0.157292659675875, 0.065895406264799, 0.082308674326412),
    (75.0, 90.0, 0.120114394661634, 0.028008930383600, 0.038029967076530),
    (90.0, 105.0,0.083889418493744, 0.030782761653475, 0.015386682792448),
    (105.0,120.0,0.093422306959018, 0.001894323794060, 0.026473157293557),
    (120.0,130.0,0.054337464251720, 0.000541235369732, 0.001142242827387),
    (130.0,140.0,0.051477597712048, 0.000405926527298, 0.000604716790970),
    (140.0,150.0,0.100095328884644, 0.0,               0.000403144527313),
    (150.0,160.0,0.027645376549122, 0.0,               0.0),
]
FIRST, LAST = 5, 5 + len(ROWS) - 1        # 5..16

assert os.path.exists(BAK), "backup missing - refusing to rebuild blind"
shutil.copy2(BAK, SRC)
wb = openpyxl.load_workbook(SRC)
ws = wb["Load Case"]
unmerge(ws)
scrub(ws, range(1, 30), range(1, 12))

title_bar(ws, "  CFR27  ·  DRIVELINE TORQUE LOAD SPECTRUM",
              "  Real telemetry — comp Jun 20 (endurance) + Jun 19 (accel).  Replaces the old Motec numbers.", "E")

# ---- table ----
for i, h in enumerate(["Bin Start\n(≥ Nm)", "Bin End\n(< Nm)", "ACCEL\n(analysed)",
                       "Endurance\n@4.61", "Endurance\n@4.20"]):
    c = ws.cell(4, 1 + i, h)
    c.font = HDR; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX
ws.row_dimensions[4].height = 30

for i, (b0, b1, acc, en, en42) in enumerate(ROWS):
    r = FIRST + i
    for col, val, fmt in ((1, b0, "0.0"), (2, b1, "0.0"),
                          (3, acc, "0.0%"), (4, en, "0.0%"), (5, en42, "0.0%")):
        c = ws.cell(r, col, val)
        c.number_format = fmt; c.border = BOX
        c.alignment = CTR
    ws.cell(r, 3).fill = FILL_KEY          # highlight the column that's actually used

# ---- short notes ----
notes = [
    "P2121 / P2127 read A4:C16 automatically — column C is the case they analyse.",
    "C = ACCEL (worst case). D/E are reference — copy one into C to analyse it instead.",
    "Endurance peaks 132 Nm, ~0% above 120 Nm → it can't really hurt the gears.",
    "Accel peaks 152 Nm, ~23% above 120 Nm → this is the fatigue case.",
]
for i, n in enumerate(notes):
    c = ws.cell(18 + i, 1, ("• " + n))
    c.font = SMALL

widths(ws, [("A", 12), ("B", 11), ("C", 13), ("D", 12), ("E", 12)])
ws.sheet_view.showGridLines = False

# ---- repoint P2121 / P2127 + extend their row formulas ----
n = len(ROWS)
for sh in ("P2121", "P2127"):
    s = wb[sh]
    s["A1"] = ArrayFormula(ref="A1:C%d" % (n + 1), text="='Load Case'!A4:C%d" % LAST)
    for r in range(12, n + 2):
        for col in range(6, 11):
            src = s.cell(11, col)
            if src.data_type == "f" and isinstance(src.value, str):
                s.cell(r, col, Translator(src.value, origin=src.coordinate)
                                 .translate_formula(s.cell(r, col).coordinate))
    print("%s -> 'Load Case'!A4:C%d" % (sh, LAST))

wb.save(SRC)
print("Rebuilt:", os.path.basename(SRC))
